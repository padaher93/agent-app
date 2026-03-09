#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
import sys
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from pypdf import PdfWriter

from agent_app_dataset.agent_workflow import append_events
from agent_app_dataset.constants import CONCEPT_LABELS
from agent_app_dataset.internal_store import InternalStore


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_sheet(path: Path, rows: list[tuple[str, float]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["C1"] = "Notes"
    for idx, (name, value) in enumerate(rows, start=2):
        ws[f"A{idx}"] = name
        ws[f"B{idx}"] = value
        ws[f"C{idx}"] = "from borrower package"
    wb.save(path)


def _write_pdf(path: Path) -> None:
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    with path.open("wb") as f:
        writer.write(f)


def _seed_analyst_note(
    *,
    store: InternalStore,
    deal_id: str,
    period_id: str,
    item_id: str,
    concept_id: str,
    case_mode: str,
    subject: str,
    note_text: str,
    memo_ready: bool,
    export_ready: bool,
) -> None:
    store.upsert_analyst_note(
        deal_id=deal_id,
        period_id=period_id,
        item_id=item_id,
        concept_id=concept_id,
        concept_maturity="review",
        trust_tier="review",
        case_mode=case_mode,
        author="seed_operator",
        subject=subject,
        note_text=note_text,
        memo_ready=memo_ready,
        export_ready=export_ready,
        metadata={
            "seeded": True,
            "source": "seed_delta_review_db",
        },
    )


def _row(
    *,
    package_id: str,
    concept_id: str,
    value: float | None,
    status: str,
    confidence: float,
    doc_id: str,
    doc_name: str,
    page_or_sheet: str,
    locator_type: str,
    locator_value: str,
    snippet: str,
    blockers: list[str] | None = None,
    source_anchors: list[dict] | None = None,
    requirement_anchor: dict | None = None,
    extraction_reason_code: str | None = None,
    extraction_reason_label: str | None = None,
    extraction_reason_detail: str | None = None,
    uncertainty_source: str | None = None,
    match_basis: str | None = None,
    source_modality: str | None = None,
    candidate_count: int | None = None,
    expected_section_state: str | None = None,
) -> dict:
    trace_id = f"tr_{package_id}_{concept_id}"
    return {
        "concept_id": concept_id,
        "label": CONCEPT_LABELS.get(concept_id, concept_id),
        "status": status,
        "dictionary_version": "v1.0",
        "raw_value_text": "" if value is None else str(value),
        "normalized_value": value,
        "current_value": value,
        "unit_currency": "USD",
        "confidence": confidence,
        "hard_blockers": blockers or [],
        "trace_id": trace_id,
        "extraction_reason_code": extraction_reason_code,
        "extraction_reason_label": extraction_reason_label,
        "extraction_reason_detail": extraction_reason_detail,
        "uncertainty_source": uncertainty_source,
        "match_basis": match_basis,
        "source_modality": source_modality,
        "candidate_count": int(candidate_count or 0),
        "expected_section_state": expected_section_state,
        "source_anchors": source_anchors or [],
        "requirement_anchor": requirement_anchor,
        "evidence_link": {
            "doc_id": doc_id,
            "doc_name": doc_name,
            "page_or_sheet": page_or_sheet,
            "locator_type": locator_type,
            "locator_value": locator_value,
        },
        "evidence": {
            "doc_id": doc_id,
            "doc_name": doc_name,
            "page_or_sheet": page_or_sheet,
            "locator_type": locator_type,
            "locator_value": locator_value,
            "source_snippet": snippet,
            "raw_value_text": "" if value is None else str(value),
            "normalized_value": value,
            "unit_currency": "USD",
            "extractor_agent_id": "agent_3",
            "verifier_agent_id": "agent_4",
            "trace_id": trace_id,
            "extracted_at": _utc_now(),
        },
    }


def _manifest(
    *,
    workspace_id: str,
    package_id: str,
    deal_id: str,
    period_end_date: str,
    source_email_id: str,
    received_at: str,
    xlsx_file_id: str,
    xlsx_name: str,
    xlsx_uri: str,
    pdf_file_id: str,
    pdf_name: str,
    pdf_uri: str,
) -> dict:
    return {
        "schema_version": "1.0",
        "package_id": package_id,
        "workspace_id": workspace_id,
        "deal_id": deal_id,
        "period_end_date": period_end_date,
        "source_email_id": source_email_id,
        "received_at": received_at,
        "files": [
            {
                "file_id": xlsx_file_id,
                "source_id": f"src_{xlsx_file_id}",
                "doc_type": "XLSX",
                "filename": xlsx_name,
                "storage_uri": xlsx_uri,
                "checksum": f"checksum_{xlsx_file_id}",
                "pages_or_sheets": 1,
            },
            {
                "file_id": pdf_file_id,
                "source_id": f"src_{pdf_file_id}",
                "doc_type": "PDF",
                "filename": pdf_name,
                "storage_uri": pdf_uri,
                "checksum": f"checksum_{pdf_file_id}",
                "pages_or_sheets": 1,
            },
        ],
        "source_ids": [f"src_{xlsx_file_id}", f"src_{pdf_file_id}"],
        "variant_tags": ["seed_delta_review"],
        "quality_flags": [],
        "labeling_workflow": {
            "primary_labeler_status": "seeded",
            "reviewer_status": "seeded",
            "adjudication_status": "not_required",
        },
        "notes": "Seeded for Delta Review v2 surface.",
    }


def seed_delta_review(
    *,
    db_path: Path,
    events_log: Path,
    workspace_id: str,
    deal_id: str,
    deal_name: str,
    docs_dir: Path,
    include_baseline: bool = False,
    delta_showcase: bool = False,
) -> dict:
    docs_dir.mkdir(parents=True, exist_ok=True)
    baseline_xlsx = docs_dir / f"{deal_id}_baseline.xlsx"
    current_xlsx = docs_dir / f"{deal_id}_current.xlsx"
    current_memo_xlsx = docs_dir / f"{deal_id}_current_memo.xlsx"
    baseline_pdf = docs_dir / f"{deal_id}_baseline.pdf"
    current_pdf = docs_dir / f"{deal_id}_current.pdf"
    reporting_pdf = docs_dir / f"{deal_id}_reporting_requirements.pdf"

    _write_sheet(
        baseline_xlsx,
        [
            ("Revenue (Total)", 12_500_000),
            ("EBITDA (Reported)", 2_580_000),
            ("EBITDA (Adjusted)", 2_610_000),
            ("Net Income", 980_000),
            ("Cash and Equivalents", 1_480_000),
            ("Total Debt", 10_600_000),
            ("Accounts Receivable (Total)", 2_740_000),
        ],
    )
    _write_sheet(
        current_xlsx,
        [
            ("Revenue (Total)", 12_450_000),
            ("EBITDA (Reported)", 2_440_000),
            ("EBITDA (Adjusted)", 2_450_000),
            ("Net Income", 0),
            ("Cash and Equivalents", 0),
            ("Total Debt", 11_200_000),
            ("Accounts Receivable (Total)", 2_680_000),
        ],
    )
    _write_sheet(
        current_memo_xlsx,
        [
            ("Revenue (Total)", 12_150_000),
            ("EBITDA (Reported)", 2_460_000),
            ("EBITDA (Adjusted)", 2_450_000),
            ("Net Income", 960_000),
        ],
    )
    _write_pdf(baseline_pdf)
    _write_pdf(current_pdf)
    _write_pdf(reporting_pdf)

    store = InternalStore(db_path)
    store.ensure_deal(deal_id=deal_id, display_name=deal_name)
    store.assign_deal_workspace(deal_id=deal_id, workspace_id=workspace_id)

    baseline_package_id = f"{deal_id}_period_2025_06_30"
    current_package_id = f"{deal_id}_period_2025_09_30"
    baseline_received_at = "2025-07-05T12:00:00+00:00"
    current_received_at = "2025-10-05T12:00:00+00:00"

    baseline_manifest = _manifest(
        workspace_id=workspace_id,
        package_id=baseline_package_id,
        deal_id=deal_id,
        period_end_date="2025-06-30",
        source_email_id="seed_email_baseline",
        received_at=baseline_received_at,
        xlsx_file_id="file_seed_baseline_xlsx",
        xlsx_name=baseline_xlsx.name,
        xlsx_uri=str(baseline_xlsx.resolve()),
        pdf_file_id="file_seed_baseline_pdf",
        pdf_name=baseline_pdf.name,
        pdf_uri=str(baseline_pdf.resolve()),
    )
    current_manifest = _manifest(
        workspace_id=workspace_id,
        package_id=current_package_id,
        deal_id=deal_id,
        period_end_date="2025-09-30",
        source_email_id="seed_email_current",
        received_at=current_received_at,
        xlsx_file_id="file_seed_current_xlsx",
        xlsx_name=current_xlsx.name,
        xlsx_uri=str(current_xlsx.resolve()),
        pdf_file_id="file_seed_current_pdf",
        pdf_name=current_pdf.name,
        pdf_uri=str(current_pdf.resolve()),
    )
    current_manifest["files"].append(
        {
            "file_id": "file_seed_current_memo_xlsx",
            "source_id": "src_file_seed_current_memo_xlsx",
            "doc_type": "XLSX",
            "filename": current_memo_xlsx.name,
            "storage_uri": str(current_memo_xlsx.resolve()),
            "checksum": "checksum_file_seed_current_memo_xlsx",
            "pages_or_sheets": 1,
        }
    )
    current_manifest["source_ids"].append("src_file_seed_current_memo_xlsx")
    current_manifest["files"].append(
        {
            "file_id": "file_seed_reporting_pdf",
            "source_id": "src_file_seed_reporting_pdf",
            "doc_type": "PDF",
            "filename": reporting_pdf.name,
            "storage_uri": str(reporting_pdf.resolve()),
            "checksum": "checksum_file_seed_reporting_pdf",
            "pages_or_sheets": 1,
        }
    )
    current_manifest["source_ids"].append("src_file_seed_reporting_pdf")

    baseline_created = False
    if include_baseline:
        _, baseline_created = store.upsert_package(
            package_id=baseline_package_id,
            idempotency_key=f"idemp_{baseline_package_id}",
            sender_email="seed@patrici.us",
            source_email_id="seed_email_baseline",
            deal_id=deal_id,
            period_end_date="2025-06-30",
            received_at=baseline_received_at,
            status="received",
            package_manifest=baseline_manifest,
        )
    _, current_created = store.upsert_package(
        package_id=current_package_id,
        idempotency_key=f"idemp_{current_package_id}",
        sender_email="seed@patrici.us",
        source_email_id="seed_email_current",
        deal_id=deal_id,
        period_end_date="2025-09-30",
        received_at=current_received_at,
        status="received",
        package_manifest=current_manifest,
    )

    baseline_rows = [
        _row(
            package_id=baseline_package_id,
            concept_id="revenue_total",
            value=12_500_000,
            status="verified",
            confidence=0.99,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B2",
            snippet="Revenue total from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="ebitda_reported",
            value=2_580_000,
            status="verified",
            confidence=0.98,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B3",
            snippet="EBITDA reported from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="ebitda_adjusted",
            value=2_610_000,
            status="verified",
            confidence=0.99,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B4",
            snippet="EBITDA adjusted from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="net_income",
            value=980_000,
            status="verified",
            confidence=0.99,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B5",
            snippet="Net income from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="cash_and_equivalents",
            value=1_480_000,
            status="verified",
            confidence=0.99,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B6",
            snippet="Cash and equivalents from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="total_debt",
            value=10_600_000,
            status="verified",
            confidence=0.98,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B7",
            snippet="Total debt from borrower package.",
        ),
        _row(
            package_id=baseline_package_id,
            concept_id="accounts_receivable_total",
            value=2_740_000,
            status="verified",
            confidence=0.97,
            doc_id="file_seed_baseline_xlsx",
            doc_name=baseline_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B8",
            snippet="Accounts receivable from borrower package.",
        ),
    ]

    if include_baseline and delta_showcase:
        ebitda_reported_row = _row(
            package_id=current_package_id,
            concept_id="ebitda_reported",
            value=2_580_000,
            status="verified",
            confidence=0.96,
            doc_id="file_seed_current_xlsx",
            doc_name=current_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B3",
            snippet="EBITDA reported 2,580,000 from structured coverage row.",
            source_anchors=[
                {
                    "anchor_id": f"tr_{current_package_id}_ebitda_reported:cand:1",
                    "doc_id": "file_seed_current_xlsx",
                    "doc_name": current_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B3",
                    "source_snippet": "EBITDA reported 2,580,000 from structured coverage row.",
                    "raw_value_text": "2580000",
                    "normalized_value": 2580000,
                    "concept_id": "ebitda_reported",
                    "concept_label": "EBITDA (Reported)",
                    "period_id": current_package_id,
                    "trace_id": f"tr_{current_package_id}_ebitda_reported",
                    "source_role": "coverage_sheet",
                    "confidence": 0.96,
                }
            ],
        )
    else:
        ebitda_reported_row = _row(
            package_id=current_package_id,
            concept_id="ebitda_reported",
            value=2_440_000,
            status="candidate_flagged",
            confidence=0.83,
            doc_id="file_seed_current_pdf",
            doc_name=current_pdf.name,
            page_or_sheet="Page 1",
            locator_type="paragraph",
            locator_value="p1:l8",
            snippet="Management discussion cites EBITDA reported at 2,440,000.",
            extraction_reason_code="candidate_from_pdf_text_only",
            extraction_reason_label="Extracted from PDF text only",
            extraction_reason_detail="Candidate was found in PDF text, not in a structured table row.",
            uncertainty_source="package_extraction",
            match_basis="llm_candidate_match",
            source_modality="pdf_text",
            candidate_count=1,
            expected_section_state="structured_anchor_not_found",
            source_anchors=[
                {
                    "anchor_id": f"tr_{current_package_id}_ebitda_reported:cand:1",
                    "doc_id": "file_seed_current_pdf",
                    "doc_name": current_pdf.name,
                    "page_or_sheet": "Page 1",
                    "locator_type": "paragraph",
                    "locator_value": "p1:l8",
                    "source_snippet": "Management discussion cites EBITDA reported at 2,440,000.",
                    "raw_value_text": "2440000",
                    "normalized_value": 2440000,
                    "concept_id": "ebitda_reported",
                    "concept_label": "EBITDA (Reported)",
                    "period_id": current_package_id,
                    "trace_id": f"tr_{current_package_id}_ebitda_reported",
                    "source_role": "submitted_source_line",
                    "confidence": 0.83,
                }
            ],
        )

    if include_baseline and delta_showcase:
        current_rows = [
            _row(
                package_id=current_package_id,
                concept_id="revenue_total",
                value=12_460_000,
                status="verified",
                confidence=0.97,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B2",
                snippet="Revenue total 12,460,000 from submitted coverage sheet.",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_revenue_total:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B2",
                        "source_snippet": "Revenue total 12,460,000 from submitted coverage sheet.",
                        "raw_value_text": "12460000",
                        "normalized_value": 12460000,
                        "concept_id": "revenue_total",
                        "concept_label": "Revenue (Total)",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_revenue_total",
                        "source_role": "coverage_sheet",
                        "confidence": 0.97,
                    },
                ],
            ),
            _row(
                package_id=current_package_id,
                concept_id="ebitda_adjusted",
                value=2_560_000,
                status="verified",
                confidence=0.95,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B4",
                snippet="EBITDA adjusted 2,560,000 from structured coverage row.",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_ebitda_adjusted:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B4",
                        "source_snippet": "EBITDA adjusted 2,560,000 from structured coverage row.",
                        "raw_value_text": "2560000",
                        "normalized_value": 2560000,
                        "concept_id": "ebitda_adjusted",
                        "concept_label": "EBITDA (Adjusted)",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_ebitda_adjusted",
                        "source_role": "coverage_sheet",
                        "confidence": 0.95,
                    }
                ],
            ),
            ebitda_reported_row,
            _row(
                package_id=current_package_id,
                concept_id="net_income",
                value=970_000,
                status="verified",
                confidence=0.96,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B5",
                snippet="Net income 970,000 from submitted package.",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_net_income:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B5",
                        "source_snippet": "Net income 970,000 from submitted package.",
                        "raw_value_text": "970000",
                        "normalized_value": 970000,
                        "concept_id": "net_income",
                        "concept_label": "Net Income",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_net_income",
                        "source_role": "coverage_sheet",
                        "confidence": 0.96,
                    },
                ],
            ),
            _row(
                package_id=current_package_id,
                concept_id="cash_and_equivalents",
                value=1_430_000,
                status="verified",
                confidence=0.95,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B6",
                snippet="Cash and equivalents 1,430,000 from submitted package.",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_cash_and_equivalents:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B6",
                        "source_snippet": "Cash and equivalents 1,430,000 from submitted package.",
                        "raw_value_text": "1430000",
                        "normalized_value": 1430000,
                        "concept_id": "cash_and_equivalents",
                        "concept_label": "Cash and Equivalents",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_cash_and_equivalents",
                        "source_role": "coverage_sheet",
                        "confidence": 0.95,
                    }
                ],
            ),
            _row(
                package_id=current_package_id,
                concept_id="total_debt",
                value=11_200_000,
                status="candidate_flagged",
                confidence=0.88,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B7",
                snippet="Total debt from borrower package.",
            ),
            _row(
                package_id=current_package_id,
                concept_id="accounts_receivable_total",
                value=2_680_000,
                status="verified",
                confidence=0.97,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B8",
                snippet="Accounts receivable from borrower package.",
            ),
        ]
    else:
        current_rows = [
            _row(
                package_id=current_package_id,
                concept_id="revenue_total",
                value=12_450_000,
                status="unresolved",
                confidence=0.89,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B2",
                snippet="Revenue total 12,450,000 in submitted coverage sheet.",
                blockers=["currency_unit_mismatch"],
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_revenue_total:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B2",
                        "source_snippet": "Revenue total 12,450,000 in submitted coverage sheet.",
                        "raw_value_text": "12450000",
                        "normalized_value": 12450000,
                        "concept_id": "revenue_total",
                        "concept_label": "Revenue (Total)",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_revenue_total",
                        "source_role": "coverage_sheet",
                        "confidence": 0.95,
                    },
                    {
                        "anchor_id": f"tr_{current_package_id}_revenue_total:cand:2",
                        "doc_id": "file_seed_current_memo_xlsx",
                        "doc_name": current_memo_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B2",
                        "source_snippet": "Revenue total 12,150,000 in management memo workbook.",
                        "raw_value_text": "12150000",
                        "normalized_value": 12150000,
                        "concept_id": "revenue_total",
                        "concept_label": "Revenue (Total)",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_revenue_total",
                        "source_role": "management_memo",
                        "confidence": 0.94,
                    },
                ],
            ),
            _row(
                package_id=current_package_id,
                concept_id="ebitda_adjusted",
                value=2_450_000,
                status="candidate_flagged",
                confidence=0.84,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="paragraph",
                locator_value="inferred:ebitda_adjusted_row",
                snippet="Adjusted EBITDA candidate found but exact row header could not be anchored.",
                extraction_reason_code="exact_row_header_missing",
                extraction_reason_label="Exact row header missing",
                extraction_reason_detail="Candidate value was found without a precise structured row locator.",
                uncertainty_source="package_extraction",
                match_basis="label_variant_match",
                source_modality="table_cell",
                candidate_count=2,
                expected_section_state="structured_locator_missing",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_ebitda_adjusted:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B4",
                        "source_snippet": "Adjusted EBITDA (variant header) 2,450,000.",
                        "raw_value_text": "2450000",
                        "normalized_value": 2450000,
                        "concept_id": "ebitda_adjusted",
                        "concept_label": "EBITDA (Adjusted)",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_ebitda_adjusted",
                        "source_role": "coverage_sheet",
                        "confidence": 0.84,
                    }
                ],
            ),
            ebitda_reported_row,
            _row(
                package_id=current_package_id,
                concept_id="net_income",
                value=None,
                status="unresolved",
                confidence=0.21,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="paragraph",
                locator_value="unresolved:not_found",
                snippet="Net income line is missing from the submitted package.",
                blockers=["missing_source_schedule"],
                requirement_anchor={
                    "doc_id": "file_seed_reporting_pdf",
                    "doc_name": reporting_pdf.name,
                    "page_or_sheet": "Page 1",
                    "locator_type": "paragraph",
                    "locator_value": "p1:l3",
                    "source_snippet": (
                        "Quarterly reporting package must include Net Income statement "
                        "for the reporting period."
                    ),
                    "required_concept_id": "net_income",
                    "required_concept_label": "Net Income",
                    "obligation_type": "reporting_requirement",
                    "source_role": "credit_reporting_schedule",
                    "trace_id": f"tr_{current_package_id}_net_income_req",
                },
            ),
            _row(
                package_id=current_package_id,
                concept_id="cash_and_equivalents",
                value=1_090_000,
                status="candidate_flagged",
                confidence=0.82,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B6",
                snippet="Cash line candidate appears under liquidity summary.",
                source_anchors=[
                    {
                        "anchor_id": f"tr_{current_package_id}_cash_and_equivalents:cand:1",
                        "doc_id": "file_seed_current_xlsx",
                        "doc_name": current_xlsx.name,
                        "page_or_sheet": "Sheet: Coverage",
                        "locator_type": "cell",
                        "locator_value": "B6",
                        "source_snippet": "Cash and equivalents candidate 1,090,000.",
                        "raw_value_text": "1090000",
                        "normalized_value": 1090000,
                        "concept_id": "cash_and_equivalents",
                        "concept_label": "Cash and Equivalents",
                        "period_id": current_package_id,
                        "trace_id": f"tr_{current_package_id}_cash_and_equivalents",
                        "source_role": "coverage_sheet",
                        "confidence": 0.82,
                    }
                ],
            ),
            _row(
                package_id=current_package_id,
                concept_id="total_debt",
                value=11_200_000,
                status="candidate_flagged",
                confidence=0.88,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B7",
                snippet="Total debt from borrower package.",
            ),
            _row(
                package_id=current_package_id,
                concept_id="accounts_receivable_total",
                value=2_680_000,
                status="verified",
                confidence=0.97,
                doc_id="file_seed_current_xlsx",
                doc_name=current_xlsx.name,
                page_or_sheet="Sheet: Coverage",
                locator_type="cell",
                locator_value="B8",
                snippet="Accounts receivable from borrower package.",
            ),
        ]

    baseline_payload = {
        "schema_version": "1.0",
        "generator": "delta_review_seed",
        "packages": [
            {
                "package_id": baseline_package_id,
                "deal_id": deal_id,
                "period_end_date": "2025-06-30",
                "rows": baseline_rows,
            }
        ],
    }
    current_payload = {
        "schema_version": "1.0",
        "generator": "delta_review_seed",
        "packages": [
            {
                "package_id": current_package_id,
                "deal_id": deal_id,
                "period_end_date": "2025-09-30",
                "rows": current_rows,
            }
        ],
    }

    if include_baseline:
        store.update_package_status(
            package_id=baseline_package_id,
            status="completed",
            processed_payload=baseline_payload,
        )
    store.update_package_status(
        package_id=current_package_id,
        status="needs_review",
        processed_payload=current_payload,
    )
    if include_baseline:
        store.upsert_traces(
            package_id=baseline_package_id,
            deal_id=deal_id,
            period_id=baseline_package_id,
            rows=baseline_rows,
        )
    store.upsert_traces(
        package_id=current_package_id,
        deal_id=deal_id,
        period_id=current_package_id,
        rows=current_rows,
    )

    debt_trace = f"tr_{current_package_id}_total_debt"
    debt_before = next(row for row in current_rows if row["concept_id"] == "total_debt")
    debt_after = deepcopy(debt_before)
    debt_after["status"] = "verified"
    debt_after["confidence"] = 0.95
    debt_after["resolved_by_user"] = True
    debt_after["user_resolution"] = {
        "resolver": "seed_operator",
        "resolved_at": _utc_now(),
        "selected_evidence": {
            "doc_id": "file_seed_current_xlsx",
            "locator_type": "cell",
            "locator_value": "B6",
        },
        "note": "Seeded resolved debt row.",
    }

    resolution_created = False
    latest_resolution = store.get_latest_trace_resolution(debt_trace)
    if not (
        isinstance(latest_resolution, dict)
        and str(latest_resolution.get("resolver", "")).strip() == "seed_operator"
        and str(latest_resolution.get("note", "")).strip() == "Seeded resolved debt row."
    ):
        store.append_trace_resolution(
            trace_id=debt_trace,
            package_id=current_package_id,
            resolver="seed_operator",
            selected_evidence={
                "doc_id": "file_seed_current_xlsx",
                "locator_type": "cell",
                "locator_value": "B7",
            },
            note="Seeded resolved debt row.",
            row_before=debt_before,
            row_after=debt_after,
        )
        resolution_created = True
    store.update_package_status(
        package_id=current_package_id,
        status=store.compute_effective_package_status(current_package_id),
        preserve_payload=True,
    )

    if include_baseline and delta_showcase:
        _seed_analyst_note(
            store=store,
            deal_id=deal_id,
            period_id=current_package_id,
            item_id=f"rq_{current_package_id}_ebitda_adjusted",
            concept_id="ebitda_adjusted",
            case_mode="review_possible_material_change",
            subject="EBITDA variance review (Sep 2025)",
            note_text=(
                "Variance exceeds minor policy but source anchor is structured. "
                "Keep as review signal and confirm with borrower package cover sheet before escalation."
            ),
            memo_ready=True,
            export_ready=False,
        )

    events: list[dict[str, Any]] = []
    if include_baseline and baseline_created:
        events.append(
            {
                "timestamp": _utc_now(),
                "event_type": "seed_package_loaded",
                "phase": "publish",
                "agent_id": "seed",
                "package_id": baseline_package_id,
                "trace_id": "",
                "payload": {"deal_id": deal_id},
            }
        )
    if current_created:
        events.append(
            {
                "timestamp": _utc_now(),
                "event_type": "seed_package_loaded",
                "phase": "publish",
                "agent_id": "seed",
                "package_id": current_package_id,
                "trace_id": "",
                "payload": {"deal_id": deal_id},
            }
        )
    if resolution_created:
        events.append(
            {
                "timestamp": _utc_now(),
                "event_type": "user_resolved",
                "phase": "publish",
                "agent_id": "seed",
                "package_id": current_package_id,
                "trace_id": debt_trace,
                "payload": {"resolver": "seed_operator", "status": "verified"},
            }
        )
    if events:
        append_events(events_log, events)

    return {
        "deal_id": deal_id,
        "deal_name": deal_name,
        "workspace_id": workspace_id,
        "baseline_package_id": baseline_package_id if include_baseline else "",
        "current_package_id": current_package_id,
        "product_mode": "delta_review" if include_baseline else "first_package_intake",
        "docs_dir": str(docs_dir),
    }


def seed_canonical_demo_deals(
    *,
    db_path: Path,
    events_log: Path,
    workspace_id: str,
    docs_dir: Path,
) -> dict[str, dict[str, str]]:
    intake_summary = seed_delta_review(
        db_path=db_path,
        events_log=events_log,
        workspace_id=workspace_id,
        deal_id="deal_alderon",
        deal_name="Alderon Credit Partners",
        docs_dir=docs_dir,
        include_baseline=False,
    )
    delta_summary = seed_delta_review(
        db_path=db_path,
        events_log=events_log,
        workspace_id=workspace_id,
        deal_id="deal_northstar",
        deal_name="Northstar Credit Partners",
        docs_dir=docs_dir,
        include_baseline=True,
        delta_showcase=True,
    )
    return {
        "intake_demo": intake_summary,
        "delta_demo": delta_summary,
    }


def seed_northstar_followup_package(
    *,
    db_path: Path,
    events_log: Path,
    workspace_id: str,
    docs_dir: Path,
) -> dict[str, str]:
    docs_dir.mkdir(parents=True, exist_ok=True)
    deal_id = "deal_northstar"
    deal_name = "Northstar Credit Partners"
    followup_package_id = f"{deal_id}_period_2025_12_31"
    followup_received_at = "2026-01-07T12:00:00+00:00"

    followup_xlsx = docs_dir / f"{deal_id}_followup.xlsx"
    followup_memo_xlsx = docs_dir / f"{deal_id}_followup_memo.xlsx"
    followup_pdf = docs_dir / f"{deal_id}_followup.pdf"
    followup_reporting_pdf = docs_dir / f"{deal_id}_followup_reporting_requirements.pdf"

    _write_sheet(
        followup_xlsx,
        [
            ("Revenue (Total)", 12_180_000),
            ("EBITDA (Reported)", 2_530_000),
            ("EBITDA (Adjusted)", 2_390_000),
            ("Net Income", 0),
            ("Cash and Equivalents", 1_020_000),
            ("Total Debt", 11_500_000),
            ("Accounts Receivable (Total)", 2_610_000),
        ],
    )
    _write_sheet(
        followup_memo_xlsx,
        [
            ("Revenue (Total)", 11_940_000),
            ("EBITDA (Reported)", 2_500_000),
            ("EBITDA (Adjusted)", 2_380_000),
            ("Net Income", 910_000),
        ],
    )
    _write_pdf(followup_pdf)
    _write_pdf(followup_reporting_pdf)

    store = InternalStore(db_path)
    store.ensure_deal(deal_id=deal_id, display_name=deal_name)
    store.assign_deal_workspace(deal_id=deal_id, workspace_id=workspace_id)

    followup_manifest = _manifest(
        workspace_id=workspace_id,
        package_id=followup_package_id,
        deal_id=deal_id,
        period_end_date="2025-12-31",
        source_email_id="seed_email_followup",
        received_at=followup_received_at,
        xlsx_file_id="file_seed_followup_xlsx",
        xlsx_name=followup_xlsx.name,
        xlsx_uri=str(followup_xlsx.resolve()),
        pdf_file_id="file_seed_followup_pdf",
        pdf_name=followup_pdf.name,
        pdf_uri=str(followup_pdf.resolve()),
    )
    followup_manifest["files"].append(
        {
            "file_id": "file_seed_followup_memo_xlsx",
            "source_id": "src_file_seed_followup_memo_xlsx",
            "doc_type": "XLSX",
            "filename": followup_memo_xlsx.name,
            "storage_uri": str(followup_memo_xlsx.resolve()),
            "checksum": "checksum_file_seed_followup_memo_xlsx",
            "pages_or_sheets": 1,
        }
    )
    followup_manifest["source_ids"].append("src_file_seed_followup_memo_xlsx")
    followup_manifest["files"].append(
        {
            "file_id": "file_seed_followup_reporting_pdf",
            "source_id": "src_file_seed_followup_reporting_pdf",
            "doc_type": "PDF",
            "filename": followup_reporting_pdf.name,
            "storage_uri": str(followup_reporting_pdf.resolve()),
            "checksum": "checksum_file_seed_followup_reporting_pdf",
            "pages_or_sheets": 1,
        }
    )
    followup_manifest["source_ids"].append("src_file_seed_followup_reporting_pdf")

    _, package_created = store.upsert_package(
        package_id=followup_package_id,
        idempotency_key=f"idemp_{followup_package_id}",
        sender_email="seed@patrici.us",
        source_email_id="seed_email_followup",
        deal_id=deal_id,
        period_end_date="2025-12-31",
        received_at=followup_received_at,
        status="received",
        package_manifest=followup_manifest,
    )

    followup_rows = [
        _row(
            package_id=followup_package_id,
            concept_id="revenue_total",
            value=12_180_000,
            status="unresolved",
            confidence=0.87,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B2",
            snippet="Revenue total 12,180,000 in submitted follow-up coverage sheet.",
            blockers=["currency_unit_mismatch"],
            source_anchors=[
                {
                    "anchor_id": f"tr_{followup_package_id}_revenue_total:cand:1",
                    "doc_id": "file_seed_followup_xlsx",
                    "doc_name": followup_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B2",
                    "source_snippet": "Revenue total 12,180,000 in submitted follow-up coverage sheet.",
                    "raw_value_text": "12180000",
                    "normalized_value": 12180000,
                    "concept_id": "revenue_total",
                    "concept_label": "Revenue (Total)",
                    "period_id": followup_package_id,
                    "trace_id": f"tr_{followup_package_id}_revenue_total",
                    "source_role": "coverage_sheet",
                    "confidence": 0.93,
                },
                {
                    "anchor_id": f"tr_{followup_package_id}_revenue_total:cand:2",
                    "doc_id": "file_seed_followup_memo_xlsx",
                    "doc_name": followup_memo_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B2",
                    "source_snippet": "Revenue total 11,940,000 in management memo workbook.",
                    "raw_value_text": "11940000",
                    "normalized_value": 11940000,
                    "concept_id": "revenue_total",
                    "concept_label": "Revenue (Total)",
                    "period_id": followup_package_id,
                    "trace_id": f"tr_{followup_package_id}_revenue_total",
                    "source_role": "management_memo",
                    "confidence": 0.92,
                },
            ],
        ),
        _row(
            package_id=followup_package_id,
            concept_id="ebitda_adjusted",
            value=2_390_000,
            status="candidate_flagged",
            confidence=0.84,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="paragraph",
            locator_value="inferred:ebitda_adjusted_row",
            snippet="Adjusted EBITDA candidate found but exact row header could not be anchored.",
            extraction_reason_code="exact_row_header_missing",
            extraction_reason_label="Exact row header missing",
            extraction_reason_detail="Candidate value was found without a precise structured row locator.",
            uncertainty_source="package_extraction",
            match_basis="label_variant_match",
            source_modality="table_cell",
            candidate_count=2,
            expected_section_state="structured_locator_missing",
            source_anchors=[
                {
                    "anchor_id": f"tr_{followup_package_id}_ebitda_adjusted:cand:1",
                    "doc_id": "file_seed_followup_xlsx",
                    "doc_name": followup_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B4",
                    "source_snippet": "Adjusted EBITDA (variant header) 2,390,000.",
                    "raw_value_text": "2390000",
                    "normalized_value": 2390000,
                    "concept_id": "ebitda_adjusted",
                    "concept_label": "EBITDA (Adjusted)",
                    "period_id": followup_package_id,
                    "trace_id": f"tr_{followup_package_id}_ebitda_adjusted",
                    "source_role": "coverage_sheet",
                    "confidence": 0.84,
                }
            ],
        ),
        _row(
            package_id=followup_package_id,
            concept_id="ebitda_reported",
            value=2_530_000,
            status="verified",
            confidence=0.96,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B3",
            snippet="EBITDA reported 2,530,000 from structured follow-up coverage row.",
            source_anchors=[
                {
                    "anchor_id": f"tr_{followup_package_id}_ebitda_reported:cand:1",
                    "doc_id": "file_seed_followup_xlsx",
                    "doc_name": followup_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B3",
                    "source_snippet": "EBITDA reported 2,530,000 from structured follow-up coverage row.",
                    "raw_value_text": "2530000",
                    "normalized_value": 2530000,
                    "concept_id": "ebitda_reported",
                    "concept_label": "EBITDA (Reported)",
                    "period_id": followup_package_id,
                    "trace_id": f"tr_{followup_package_id}_ebitda_reported",
                    "source_role": "coverage_sheet",
                    "confidence": 0.96,
                }
            ],
        ),
        _row(
            package_id=followup_package_id,
            concept_id="net_income",
            value=None,
            status="unresolved",
            confidence=0.23,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="paragraph",
            locator_value="unresolved:not_found",
            snippet="Net income line is missing from the follow-up submitted package.",
            blockers=["missing_source_schedule"],
            requirement_anchor={
                "doc_id": "file_seed_followup_reporting_pdf",
                "doc_name": followup_reporting_pdf.name,
                "page_or_sheet": "Page 1",
                "locator_type": "paragraph",
                "locator_value": "p1:l3",
                "source_snippet": (
                    "Quarterly reporting package must include Net Income statement "
                    "for the reporting period."
                ),
                "required_concept_id": "net_income",
                "required_concept_label": "Net Income",
                "obligation_type": "reporting_requirement",
                "source_role": "credit_reporting_schedule",
                "trace_id": f"tr_{followup_package_id}_net_income_req",
            },
        ),
        _row(
            package_id=followup_package_id,
            concept_id="cash_and_equivalents",
            value=1_020_000,
            status="candidate_flagged",
            confidence=0.82,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B6",
            snippet="Cash line candidate appears under liquidity summary in follow-up package.",
            source_anchors=[
                {
                    "anchor_id": f"tr_{followup_package_id}_cash_and_equivalents:cand:1",
                    "doc_id": "file_seed_followup_xlsx",
                    "doc_name": followup_xlsx.name,
                    "page_or_sheet": "Sheet: Coverage",
                    "locator_type": "cell",
                    "locator_value": "B6",
                    "source_snippet": "Cash and equivalents candidate 1,020,000.",
                    "raw_value_text": "1020000",
                    "normalized_value": 1020000,
                    "concept_id": "cash_and_equivalents",
                    "concept_label": "Cash and Equivalents",
                    "period_id": followup_package_id,
                    "trace_id": f"tr_{followup_package_id}_cash_and_equivalents",
                    "source_role": "coverage_sheet",
                    "confidence": 0.82,
                }
            ],
        ),
        _row(
            package_id=followup_package_id,
            concept_id="total_debt",
            value=11_500_000,
            status="candidate_flagged",
            confidence=0.89,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B7",
            snippet="Total debt from follow-up borrower package.",
        ),
        _row(
            package_id=followup_package_id,
            concept_id="accounts_receivable_total",
            value=2_610_000,
            status="verified",
            confidence=0.97,
            doc_id="file_seed_followup_xlsx",
            doc_name=followup_xlsx.name,
            page_or_sheet="Sheet: Coverage",
            locator_type="cell",
            locator_value="B8",
            snippet="Accounts receivable from follow-up borrower package.",
        ),
    ]

    followup_payload = {
        "schema_version": "1.0",
        "generator": "delta_review_seed",
        "packages": [
            {
                "package_id": followup_package_id,
                "deal_id": deal_id,
                "period_end_date": "2025-12-31",
                "rows": followup_rows,
            }
        ],
    }

    store.update_package_status(
        package_id=followup_package_id,
        status="needs_review",
        processed_payload=followup_payload,
    )
    store.upsert_traces(
        package_id=followup_package_id,
        deal_id=deal_id,
        period_id=followup_package_id,
        rows=followup_rows,
    )

    debt_trace = f"tr_{followup_package_id}_total_debt"
    debt_before = next(row for row in followup_rows if row["concept_id"] == "total_debt")
    debt_after = deepcopy(debt_before)
    debt_after["status"] = "verified"
    debt_after["confidence"] = 0.95
    debt_after["resolved_by_user"] = True
    debt_after["user_resolution"] = {
        "resolver": "seed_operator",
        "resolved_at": _utc_now(),
        "selected_evidence": {
            "doc_id": "file_seed_followup_xlsx",
            "locator_type": "cell",
            "locator_value": "B7",
        },
        "note": "Seeded resolved debt row.",
    }

    resolution_created = False
    latest_resolution = store.get_latest_trace_resolution(debt_trace)
    if not (
        isinstance(latest_resolution, dict)
        and str(latest_resolution.get("resolver", "")).strip() == "seed_operator"
        and str(latest_resolution.get("note", "")).strip() == "Seeded resolved debt row."
    ):
        store.append_trace_resolution(
            trace_id=debt_trace,
            package_id=followup_package_id,
            resolver="seed_operator",
            selected_evidence={
                "doc_id": "file_seed_followup_xlsx",
                "locator_type": "cell",
                "locator_value": "B7",
            },
            note="Seeded resolved debt row.",
            row_before=debt_before,
            row_after=debt_after,
        )
        resolution_created = True

    store.update_package_status(
        package_id=followup_package_id,
        status=store.compute_effective_package_status(followup_package_id),
        preserve_payload=True,
    )

    _seed_analyst_note(
        store=store,
        deal_id=deal_id,
        period_id=followup_package_id,
        item_id=f"rq_{followup_package_id}_ebitda_adjusted",
        concept_id="ebitda_adjusted",
        case_mode="review_possible_requirement",
        subject="EBITDA support gap review (Dec 2025)",
        note_text=(
            "Adjusted EBITDA candidate is present but exact row header is missing. "
            "Treat as provisional until a structured coverage-row anchor is confirmed."
        ),
        memo_ready=True,
        export_ready=True,
    )

    events: list[dict[str, Any]] = []
    if package_created:
        events.append(
            {
                "timestamp": _utc_now(),
                "event_type": "seed_package_loaded",
                "phase": "publish",
                "agent_id": "seed",
                "package_id": followup_package_id,
                "trace_id": "",
                "payload": {"deal_id": deal_id},
            }
        )
    if resolution_created:
        events.append(
            {
                "timestamp": _utc_now(),
                "event_type": "user_resolved",
                "phase": "publish",
                "agent_id": "seed",
                "package_id": followup_package_id,
                "trace_id": debt_trace,
                "payload": {"resolver": "seed_operator", "status": "verified"},
            }
        )
    if events:
        append_events(events_log, events)

    return {
        "deal_id": deal_id,
        "deal_name": deal_name,
        "workspace_id": workspace_id,
        "baseline_package_id": "deal_northstar_period_2025_09_30",
        "current_package_id": followup_package_id,
        "product_mode": "delta_review",
        "docs_dir": str(docs_dir),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed deterministic Delta Review v2 data into InternalStore DB.")
    parser.add_argument("--db-path", default="runtime/internal_api.sqlite3")
    parser.add_argument("--events-log", default="runtime/agent_events.jsonl")
    parser.add_argument("--workspace-id", default="ws_default")
    parser.add_argument("--deal-id", default="deal_delta_review_demo")
    parser.add_argument("--deal-name", default="Alderon Credit Partners")
    parser.add_argument("--docs-dir", default="runtime/seed_docs")
    parser.add_argument("--include-baseline", action="store_true")
    parser.add_argument("--seed-canonical-demo-deals", action="store_true")
    parser.add_argument("--seed-northstar-followup", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.seed_canonical_demo_deals and args.seed_northstar_followup:
        canonical = seed_canonical_demo_deals(
            db_path=Path(args.db_path),
            events_log=Path(args.events_log),
            workspace_id=args.workspace_id,
            docs_dir=Path(args.docs_dir),
        )
        followup = seed_northstar_followup_package(
            db_path=Path(args.db_path),
            events_log=Path(args.events_log),
            workspace_id=args.workspace_id,
            docs_dir=Path(args.docs_dir),
        )
        summary = {
            **canonical,
            "northstar_followup": followup,
        }
    elif args.seed_canonical_demo_deals:
        summary = seed_canonical_demo_deals(
            db_path=Path(args.db_path),
            events_log=Path(args.events_log),
            workspace_id=args.workspace_id,
            docs_dir=Path(args.docs_dir),
        )
    elif args.seed_northstar_followup:
        summary = seed_northstar_followup_package(
            db_path=Path(args.db_path),
            events_log=Path(args.events_log),
            workspace_id=args.workspace_id,
            docs_dir=Path(args.docs_dir),
        )
    else:
        summary = seed_delta_review(
            db_path=Path(args.db_path),
            events_log=Path(args.events_log),
            workspace_id=args.workspace_id,
            deal_id=args.deal_id,
            deal_name=args.deal_name,
            docs_dir=Path(args.docs_dir),
            include_baseline=bool(args.include_baseline),
        )
    print("Seed complete:")
    if args.seed_canonical_demo_deals and args.seed_northstar_followup:
        for name, deal_summary in summary.items():
            print(f"- {name}:")
            for key, value in deal_summary.items():
                print(f"  - {key}: {value}")
    elif args.seed_canonical_demo_deals:
        for name, deal_summary in summary.items():
            print(f"- {name}:")
            for key, value in deal_summary.items():
                print(f"  - {key}: {value}")
    elif args.seed_northstar_followup:
        for key, value in summary.items():
            print(f"- {key}: {value}")
    else:
        for key, value in summary.items():
            print(f"- {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
