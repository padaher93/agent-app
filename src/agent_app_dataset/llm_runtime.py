from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import json
import os
from pathlib import Path
import re
from typing import Any, Protocol

from .constants import CONCEPT_DEFINITIONS, CONCEPT_LABELS, STARTER_CONCEPT_IDS
from .normalization import normalize_value
from .policy import classify_status
from .source_grounding import unique_trustworthy_anchors
from .storage import resolve_storage_uri

try:  # pragma: no cover - optional dependency path
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None

try:  # pragma: no cover - optional dependency path
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency path
    PdfReader = None


@dataclass(frozen=True)
class DocumentLine:
    doc_id: str
    doc_name: str
    doc_type: str
    page_or_sheet: str
    locator_type: str
    locator_value: str
    text: str


class LLMClient(Protocol):
    def run_json(
        self,
        *,
        agent_id: str,
        system_prompt: str,
        user_payload: dict[str, Any],
    ) -> dict[str, Any]: ...


class OpenAIJSONClient:
    def __init__(
        self,
        *,
        api_key: str,
        model: str = "gpt-5-mini",
        base_url: str | None = None,
    ) -> None:
        try:
            from openai import OpenAI
        except Exception as exc:  # pragma: no cover - dependency/runtime path
            raise RuntimeError("OpenAI SDK is not installed") from exc

        kwargs: dict[str, Any] = {"api_key": api_key}
        timeout_env = os.getenv("OPENAI_TIMEOUT_SECONDS")
        if timeout_env:
            try:
                kwargs["timeout"] = float(timeout_env)
            except Exception:
                kwargs["timeout"] = 60.0
        else:
            kwargs["timeout"] = 60.0
        if base_url:
            kwargs["base_url"] = base_url

        self._client = OpenAI(**kwargs)
        self._model = model

    def run_json(
        self,
        *,
        agent_id: str,
        system_prompt: str,
        user_payload: dict[str, Any],
    ) -> dict[str, Any]:
        user_text = json.dumps(user_payload, sort_keys=True)
        response = self._client.chat.completions.create(
            model=self._model,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": f"You are {agent_id}. {system_prompt} Output valid JSON only.",
                },
                {
                    "role": "user",
                    "content": user_text,
                },
            ],
        )

        choice = response.choices[0]
        raw = choice.message.content or "{}"
        if isinstance(raw, list):
            raw = "".join(part.get("text", "") for part in raw if isinstance(part, dict))

        try:
            parsed = json.loads(str(raw))
        except Exception as exc:  # pragma: no cover - runtime parse hardening
            raise RuntimeError(f"Invalid JSON response from {agent_id}") from exc

        if not isinstance(parsed, dict):
            raise RuntimeError(f"Unexpected JSON payload type from {agent_id}")
        return parsed

    @property
    def model_name(self) -> str:
        return self._model


def create_default_llm_client() -> LLMClient:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is required for extraction_mode=llm")

    model = os.getenv("PATRICIUS_LLM_MODEL", "gpt-5-mini")
    base_url = os.getenv("OPENAI_BASE_URL")
    return OpenAIJSONClient(api_key=api_key, model=model, base_url=base_url)


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_text(value: Any, max_len: int = 800) -> str:
    text = str(value or "").strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        parsed = float(value)
    except Exception:
        return default
    if parsed < 0:
        return 0.0
    if parsed > 1:
        return 1.0
    return round(parsed, 4)


def _llm_model_name(client: LLMClient) -> str:
    value = getattr(client, "model_name", None)
    if isinstance(value, str) and value.strip():
        return value.strip()
    return "unknown_model"


def _read_pdf_lines(path: Path, doc_id: str, doc_name: str) -> list[DocumentLine]:
    if PdfReader is None:
        return []

    reader = PdfReader(str(path))
    lines: list[DocumentLine] = []
    max_pages = 25

    for page_num, page in enumerate(reader.pages[:max_pages], start=1):
        text = page.extract_text() or ""
        for line_num, line in enumerate(text.splitlines(), start=1):
            clean = line.strip()
            if not clean:
                continue
            lines.append(
                DocumentLine(
                    doc_id=doc_id,
                    doc_name=doc_name,
                    doc_type="PDF",
                    page_or_sheet=f"Page {page_num}",
                    locator_type="paragraph",
                    locator_value=f"p{page_num}:l{line_num}",
                    text=_safe_text(clean, max_len=600),
                )
            )
            if len(lines) >= 1200:
                return lines

    return lines


def _extract_numeric_text(value: Any) -> str | None:
    if isinstance(value, (int, float)):
        return str(value)
    if not isinstance(value, str):
        return None

    m = re.search(r"-?[0-9][0-9,]*(?:\.[0-9]+)?", value)
    if not m:
        return None
    return m.group(0)


def _source_role_for_doc_type(doc_type: str) -> str:
    if str(doc_type).strip().upper() == "XLSX":
        return "worksheet_value"
    return "submitted_source_line"


def _source_modality_from_locator(locator_type: str, doc_type: str) -> str:
    normalized_locator = str(locator_type or "").strip().lower()
    normalized_doc_type = str(doc_type or "").strip().upper()
    if normalized_locator == "cell":
        return "table_cell"
    if normalized_doc_type == "PDF":
        return "pdf_text"
    return "narrative_text"


def _reason_payload(
    *,
    code: str | None,
    label: str | None,
    detail: str | None,
    uncertainty_source: str | None,
    match_basis: str | None,
    source_modality: str | None,
    candidate_count: int,
    expected_section_state: str | None = None,
) -> dict[str, Any]:
    return {
        "extraction_reason_code": code,
        "extraction_reason_label": label,
        "extraction_reason_detail": detail,
        "uncertainty_source": uncertainty_source,
        "match_basis": match_basis,
        "source_modality": source_modality,
        "candidate_count": int(max(candidate_count, 0)),
        "expected_section_state": expected_section_state,
    }


def _select_candidate_line(candidate: dict[str, Any], candidate_lines: list[dict[str, Any]]) -> dict[str, Any] | None:
    candidate_doc_id = str(candidate.get("doc_id", "")).strip()
    candidate_locator = str(candidate.get("locator_value", "")).strip()
    candidate_page = str(candidate.get("page_or_sheet", "")).strip()
    candidate_locator_type = str(candidate.get("locator_type", "")).strip().lower()

    for line in candidate_lines:
        line_doc_id = str(line.get("doc_id", "")).strip()
        line_locator = str(line.get("locator_value", "")).strip()
        line_page = str(line.get("page_or_sheet", "")).strip()
        line_locator_type = str(line.get("locator_type", "")).strip().lower()
        if candidate_doc_id and candidate_doc_id != line_doc_id:
            continue
        if candidate_locator and candidate_locator != line_locator:
            continue
        if candidate_page and candidate_page != line_page:
            continue
        if candidate_locator_type and candidate_locator_type != line_locator_type:
            continue
        return line

    for line in candidate_lines:
        line_doc_id = str(line.get("doc_id", "")).strip()
        if candidate_doc_id and candidate_doc_id != line_doc_id:
            continue
        return line

    return None


def _keyword_candidate_lines(candidate_lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [line for line in candidate_lines if bool(line.get("keyword_match"))]


def _line_modality(line: dict[str, Any]) -> str:
    return _source_modality_from_locator(
        locator_type=str(line.get("locator_type", "")),
        doc_type=str(line.get("doc_type", "")),
    )


def _llm_reason_for_row(
    *,
    selected_candidate: dict[str, Any] | None,
    selected_line: dict[str, Any] | None,
    candidate_lines: list[dict[str, Any]],
    status: str,
    missing_sources: bool,
    normalization_unresolved_reason: str | None,
) -> dict[str, Any]:
    reason_code: str | None = None
    reason_label: str | None = None
    reason_detail: str | None = None
    uncertainty_source: str | None = None
    expected_section_state: str | None = None

    keyword_lines = _keyword_candidate_lines(candidate_lines)
    candidate_count = len(keyword_lines)
    modalities = {_line_modality(line) for line in keyword_lines}
    modality_for_context = "mixed" if len(modalities) > 1 else (next(iter(modalities), "none"))
    match_basis = "llm_candidate_search"

    if selected_candidate is None:
        if missing_sources:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "Source files were unavailable during extraction."
            uncertainty_source = "package_extraction"
            expected_section_state = "source_document_unavailable"
        elif candidate_count == 0:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "No exact source anchor was found in the current package."
            uncertainty_source = "package_extraction"
            expected_section_state = "exact_anchor_not_found"
        elif modality_for_context == "pdf_text":
            reason_code = "candidate_from_pdf_text_only"
            reason_label = "Extracted from PDF text only"
            reason_detail = "Candidates were found in PDF text, not in structured table rows."
            uncertainty_source = "package_extraction"
            expected_section_state = "structured_anchor_not_found"
        elif modality_for_context == "narrative_text":
            reason_code = "candidate_from_narrative_text"
            reason_label = "Candidate extracted from narrative text"
            reason_detail = "Candidates were found in narrative text, not in structured table rows."
            uncertainty_source = "package_extraction"
            expected_section_state = "structured_anchor_not_found"
        elif candidate_count >= 2:
            reason_code = "multiple_matching_rows"
            reason_label = "Multiple matching rows"
            reason_detail = "More than one matching anchor was found for this concept."
            uncertainty_source = "package_extraction"
            expected_section_state = "multiple_candidate_anchors"
        return _reason_payload(
            code=reason_code,
            label=reason_label,
            detail=reason_detail,
            uncertainty_source=uncertainty_source,
            match_basis=match_basis,
            source_modality=modality_for_context,
            candidate_count=candidate_count,
            expected_section_state=expected_section_state,
        )

    candidate_locator_value = str(selected_candidate.get("locator_value", "")).strip()
    selected_locator_type = str(selected_candidate.get("locator_type", "")).strip()
    selected_doc_type = str((selected_line or {}).get("doc_type", "")).strip()
    if not selected_doc_type:
        selected_doc_name = str(selected_candidate.get("doc_name", "")).strip().lower()
        if selected_doc_name.endswith(".pdf"):
            selected_doc_type = "PDF"
    source_modality = _source_modality_from_locator(selected_locator_type, selected_doc_type)
    match_basis = "llm_verified_candidate" if status == "verified" else "llm_candidate_match"

    if status in {"candidate_flagged", "unresolved"}:
        if not candidate_locator_value or candidate_locator_value.startswith("inferred:"):
            reason_code = "exact_row_header_missing"
            reason_label = "Exact row header missing"
            reason_detail = "Candidate value was found without a precise structured row locator."
            uncertainty_source = "package_extraction"
            expected_section_state = "structured_locator_missing"
        elif source_modality == "pdf_text":
            reason_code = "candidate_from_pdf_text_only"
            reason_label = "Extracted from PDF text only"
            reason_detail = "Candidate was found in PDF text, not in a structured table row."
            uncertainty_source = "package_extraction"
        elif source_modality == "narrative_text":
            reason_code = "candidate_from_narrative_text"
            reason_label = "Candidate extracted from narrative text"
            reason_detail = "Candidate came from narrative text rather than a structured table row."
            uncertainty_source = "package_extraction"
        elif candidate_count >= 2:
            reason_code = "multiple_matching_rows"
            reason_label = "Multiple matching rows"
            reason_detail = "More than one matching anchor was found for this concept."
            uncertainty_source = "package_extraction"
        elif normalization_unresolved_reason:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "Extracted candidate could not be normalized into a trustworthy numeric value."
            uncertainty_source = "package_extraction"

    return _reason_payload(
        code=reason_code,
        label=reason_label,
        detail=reason_detail,
        uncertainty_source=uncertainty_source,
        match_basis=match_basis,
        source_modality=source_modality,
        candidate_count=candidate_count,
        expected_section_state=expected_section_state,
    )


def _line_anchor_candidates(
    *,
    concept_id: str,
    concept_label: str,
    package_id: str,
    trace_id: str,
    deal_currency: str,
    candidate_lines: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    anchors: list[dict[str, Any]] = []
    for index, line in enumerate(candidate_lines, start=1):
        line_text = str(line.get("text", "")).strip()
        raw_value_text = _extract_numeric_text(line_text) or ""
        normalized = normalize_value(
            raw_value_text=raw_value_text,
            source_snippet=line_text,
            deal_currency=deal_currency,
        )
        anchors.append(
            {
                "anchor_id": f"{trace_id}:line:{index}",
                "doc_id": str(line.get("doc_id", "")).strip(),
                "doc_name": str(line.get("doc_name", "")).strip(),
                "page_or_sheet": str(line.get("page_or_sheet", "")).strip(),
                "locator_type": str(line.get("locator_type", "")).strip(),
                "locator_value": str(line.get("locator_value", "")).strip(),
                "source_snippet": line_text,
                "raw_value_text": raw_value_text,
                "normalized_value": normalized.normalized_value,
                "unit_currency": normalized.unit_currency,
                "concept_id": concept_id,
                "concept_label": concept_label,
                "period_id": package_id,
                "trace_id": trace_id,
                "source_role": _source_role_for_doc_type(str(line.get("doc_type", ""))),
                "confidence": None,
            }
        )
    return anchors


def _read_xlsx_lines(path: Path, doc_id: str, doc_name: str) -> list[DocumentLine]:
    if load_workbook is None:
        return []

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    lines: list[DocumentLine] = []

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=500):
                row_values: list[str] = []
                first_coordinate = "A1"
                for idx, cell in enumerate(row):
                    value = cell.value
                    if idx == 0:
                        first_coordinate = getattr(cell, "coordinate", "A1")
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    row_values.append(text)

                if not row_values:
                    continue

                joined = " | ".join(row_values)
                lines.append(
                    DocumentLine(
                        doc_id=doc_id,
                        doc_name=doc_name,
                        doc_type="XLSX",
                        page_or_sheet=f"Sheet: {sheet.title}",
                        locator_type="cell",
                        locator_value=first_coordinate,
                        text=_safe_text(joined, max_len=600),
                    )
                )

                if len(lines) >= 1200:
                    return lines
    finally:
        workbook.close()

    return lines


def _collect_document_context(package_manifest: dict[str, Any]) -> tuple[list[dict[str, Any]], bool]:
    docs: list[dict[str, Any]] = []
    missing_sources = False

    for file_meta in package_manifest.get("files", []):
        doc_id = str(file_meta.get("file_id", ""))
        doc_name = str(file_meta.get("filename", ""))
        doc_type = str(file_meta.get("doc_type", "")).upper()
        storage_uri = str(file_meta.get("storage_uri", ""))
        path = resolve_storage_uri(storage_uri)

        if path is None:
            missing_sources = True
            docs.append(
                {
                    "doc_id": doc_id,
                    "doc_name": doc_name,
                    "doc_type": doc_type,
                    "storage_uri": storage_uri,
                    "available": False,
                    "lines": [],
                }
            )
            continue

        if doc_type == "PDF":
            line_objs = _read_pdf_lines(path=path, doc_id=doc_id, doc_name=doc_name)
        elif doc_type == "XLSX":
            line_objs = _read_xlsx_lines(path=path, doc_id=doc_id, doc_name=doc_name)
        else:
            line_objs = []

        docs.append(
            {
                "doc_id": doc_id,
                "doc_name": doc_name,
                "doc_type": doc_type,
                "storage_uri": storage_uri,
                "available": True,
                "line_count": len(line_objs),
                "lines": [
                    {
                        "doc_id": item.doc_id,
                        "doc_name": item.doc_name,
                        "doc_type": item.doc_type,
                        "page_or_sheet": item.page_or_sheet,
                        "locator_type": item.locator_type,
                        "locator_value": item.locator_value,
                        "text": item.text,
                    }
                    for item in line_objs
                ],
            }
        )

    return docs, missing_sources


def _pick_line_candidates(documents: list[dict[str, Any]], concept_id: str, max_items: int = 160) -> list[dict[str, Any]]:
    keywords = [str(value).lower() for value in CONCEPT_DEFINITIONS[concept_id]["keywords"]]

    candidates: list[dict[str, Any]] = []
    fallback: list[dict[str, Any]] = []
    for doc in documents:
        for line in doc.get("lines", []):
            text = str(line.get("text", ""))
            lowered = text.lower()

            line_payload = {
                "doc_id": doc.get("doc_id", ""),
                "doc_name": doc.get("doc_name", ""),
                "doc_type": doc.get("doc_type", ""),
                "page_or_sheet": line.get("page_or_sheet", ""),
                "locator_type": line.get("locator_type", "paragraph"),
                "locator_value": line.get("locator_value", ""),
                "text": _safe_text(text, max_len=320),
            }

            if any(keyword in lowered for keyword in keywords):
                line_payload["keyword_match"] = True
                candidates.append(line_payload)
            elif len(fallback) < 120:
                line_payload["keyword_match"] = False
                fallback.append(line_payload)

    merged = candidates[:max_items]
    if len(merged) < max_items:
        merged.extend(fallback[: max_items - len(merged)])
    return merged[:max_items]


def _agent2_classify_files(llm_client: LLMClient, documents: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    payload = {
        "task": "Classify each document by financial relevance for private credit package extraction.",
        "files": [
            {
                "doc_id": doc.get("doc_id", ""),
                "doc_name": doc.get("doc_name", ""),
                "doc_type": doc.get("doc_type", ""),
                "line_count": doc.get("line_count", 0),
                "samples": [line.get("text", "") for line in doc.get("lines", [])[:15]],
            }
            for doc in documents
        ],
        "allowed_labels": [
            "financial_statement",
            "compliance_certificate",
            "borrowing_base",
            "covenant_report",
            "other",
        ],
    }

    output = llm_client.run_json(
        agent_id="agent_2",
        system_prompt=(
            "You are Agent 2 (Package Understanding). "
            "Classify files and return a concise JSON array under key 'files'. "
            "Each file item must include doc_id, classification, confidence, rationale."
        ),
        user_payload=payload,
    )

    by_doc: dict[str, dict[str, Any]] = {}
    for item in output.get("files", []):
        doc_id = str(item.get("doc_id", "")).strip()
        if not doc_id:
            continue
        by_doc[doc_id] = {
            "classification": str(item.get("classification", "other")),
            "confidence": _safe_float(item.get("confidence", 0.0), default=0.0),
            "rationale": _safe_text(item.get("rationale", ""), max_len=300),
        }

    return by_doc


def _agent3_extract_candidate(
    *,
    llm_client: LLMClient,
    concept_id: str,
    concept_label: str,
    concept_keywords: list[str],
    candidate_lines: list[dict[str, Any]],
    file_classifications: dict[str, dict[str, Any]],
    prior_feedback: list[str],
) -> dict[str, Any]:
    payload = {
        "task": "Find the best candidate evidence row for the requested concept.",
        "concept": {
            "concept_id": concept_id,
            "label": concept_label,
            "keywords": concept_keywords,
        },
        "file_classifications": file_classifications,
        "candidate_lines": candidate_lines,
        "prior_feedback": prior_feedback,
        "output_contract": {
            "candidate": {
                "found": "bool",
                "raw_value_text": "string",
                "doc_id": "string",
                "doc_name": "string",
                "page_or_sheet": "string",
                "locator_type": "cell|paragraph",
                "locator_value": "string",
                "source_snippet": "string",
                "confidence": "0..1",
                "reason": "string",
            }
        },
    }

    output = llm_client.run_json(
        agent_id="agent_3",
        system_prompt=(
            "You are Agent 3 (Extractor). Select one best candidate for the concept. "
            "If no reliable candidate exists, return candidate.found=false. "
            "Use exact locator references from the provided candidate lines."
        ),
        user_payload=payload,
    )

    return dict(output.get("candidate", {}))


def _agent4_verify_candidate(
    *,
    llm_client: LLMClient,
    concept_id: str,
    concept_label: str,
    candidate: dict[str, Any],
    candidate_lines: list[dict[str, Any]],
    file_classifications: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    payload = {
        "task": "Validate if candidate evidence is appropriate for concept and should be accepted.",
        "concept": {
            "concept_id": concept_id,
            "label": concept_label,
        },
        "candidate": candidate,
        "candidate_lines": candidate_lines,
        "file_classifications": file_classifications,
        "output_contract": {
            "decision": "accept|reject",
            "confidence_adjustment": "-0.3..0.3",
            "objections": "string[]",
            "reason": "string",
        },
    }

    output = llm_client.run_json(
        agent_id="agent_4",
        system_prompt=(
            "You are Agent 4 (Independent Verifier). "
            "Reject if candidate file type/context is likely wrong or evidence is weak."
        ),
        user_payload=payload,
    )

    decision = str(output.get("decision", "reject")).strip().lower()
    if decision not in {"accept", "reject"}:
        decision = "reject"

    return {
        "decision": decision,
        "confidence_adjustment": float(output.get("confidence_adjustment", 0.0) or 0.0),
        "objections": [str(item) for item in output.get("objections", [])],
        "reason": _safe_text(output.get("reason", ""), max_len=300),
    }


def run_llm_multi_agent_extraction(
    package_manifest: dict[str, Any],
    *,
    max_retries: int = 2,
    deal_currency: str = "USD",
    llm_client: LLMClient | None = None,
) -> dict[str, Any]:
    client = llm_client or create_default_llm_client()

    package_id = str(package_manifest["package_id"])
    files = package_manifest.get("files", [])
    fallback_file = files[0] if files else {}
    fallback_doc_id = str(fallback_file.get("file_id", ""))
    fallback_doc_name = str(fallback_file.get("filename", ""))
    fallback_doc_type = str(fallback_file.get("doc_type", "")).upper()
    fallback_page = "Sheet: unknown" if fallback_doc_type == "XLSX" else "Page 1"

    extracted_at = _now_utc()
    model_name = _llm_model_name(client)
    documents, missing_sources = _collect_document_context(package_manifest)
    file_classifications = _agent2_classify_files(client, documents) if documents else {}

    rows: list[dict[str, Any]] = []

    for concept_id in STARTER_CONCEPT_IDS:
        trace_id = f"tr_{package_id}_{concept_id}"
        concept_meta = CONCEPT_DEFINITIONS[concept_id]
        concept_label = CONCEPT_LABELS[concept_id]
        concept_keywords = list(concept_meta["keywords"])

        candidate_lines = _pick_line_candidates(documents, concept_id)
        source_anchors = _line_anchor_candidates(
            concept_id=concept_id,
            concept_label=concept_label,
            package_id=package_id,
            trace_id=trace_id,
            deal_currency=deal_currency,
            candidate_lines=candidate_lines,
        )
        prior_feedback: list[str] = []
        accepted_candidate: dict[str, Any] | None = None
        final_verifier: dict[str, Any] | None = None
        unresolved_reason = "no_candidate_found"
        attempt_logs: list[dict[str, Any]] = []

        for attempt_no in range(max_retries + 1):
            candidate = _agent3_extract_candidate(
                llm_client=client,
                concept_id=concept_id,
                concept_label=concept_label,
                concept_keywords=concept_keywords,
                candidate_lines=candidate_lines,
                file_classifications=file_classifications,
                prior_feedback=prior_feedback,
            )

            found = bool(candidate.get("found"))
            if not found:
                unresolved_reason = str(candidate.get("reason", "no_candidate_found"))
                attempt_logs.append(
                    {
                        "attempt": attempt_no,
                        "agent_3_candidate": {
                            "found": False,
                            "reason": unresolved_reason,
                        },
                        "agent_4_verification": None,
                    }
                )
                break

            verifier = _agent4_verify_candidate(
                llm_client=client,
                concept_id=concept_id,
                concept_label=concept_label,
                candidate=candidate,
                candidate_lines=candidate_lines,
                file_classifications=file_classifications,
            )

            final_verifier = verifier
            attempt_logs.append(
                {
                    "attempt": attempt_no,
                    "agent_3_candidate": {
                        "found": True,
                        "confidence": _safe_float(candidate.get("confidence", 0.0), default=0.0),
                        "doc_id": str(candidate.get("doc_id", "")),
                        "page_or_sheet": str(candidate.get("page_or_sheet", "")),
                        "locator_value": str(candidate.get("locator_value", "")),
                        "reason": str(candidate.get("reason", "")),
                    },
                    "agent_4_verification": verifier,
                }
            )
            if verifier["decision"] == "accept":
                accepted_candidate = candidate
                break

            objections = verifier.get("objections", [])
            unresolved_reason = str(verifier.get("reason") or "verification_rejected")
            if objections:
                prior_feedback.extend(objections)
            else:
                prior_feedback.append(unresolved_reason)

        if accepted_candidate is None:
            blockers = ["no_reliable_candidate", unresolved_reason]
            if final_verifier and final_verifier.get("objections"):
                blockers.extend([str(item) for item in final_verifier.get("objections", [])])
            if missing_sources:
                blockers.append("missing_source_document")

            locator_value = "unresolved:not_found"
            if not fallback_doc_id:
                locator_value = "unresolved:missing_package_file"

            reason = _llm_reason_for_row(
                selected_candidate=None,
                selected_line=None,
                candidate_lines=candidate_lines,
                status="unresolved",
                missing_sources=missing_sources,
                normalization_unresolved_reason=None,
            )
            row = {
                "concept_id": concept_id,
                "label": concept_label,
                "status": "unresolved",
                "dictionary_version": "v1.0",
                "raw_value_text": "",
                "normalized_value": None,
                "current_value": None,
                "unit_currency": deal_currency,
                "confidence": 0.0,
                "hard_blockers": sorted(set(blockers)),
                "trace_id": trace_id,
                **reason,
                "source_anchors": unique_trustworthy_anchors(source_anchors, max_items=8),
                "llm_trace": {
                    "model": model_name,
                    "agent_2_classification": file_classifications,
                    "attempts": attempt_logs,
                    "final_reason": unresolved_reason,
                    "max_retries": max_retries,
                },
                "evidence_link": {
                    "doc_id": fallback_doc_id,
                    "doc_name": fallback_doc_name,
                    "page_or_sheet": fallback_page if fallback_doc_id else "Package Context",
                    "locator_type": "paragraph",
                    "locator_value": locator_value,
                },
                "evidence": {
                    "doc_id": fallback_doc_id,
                    "doc_name": fallback_doc_name,
                    "page_or_sheet": fallback_page if fallback_doc_id else "Package Context",
                    "locator_type": "paragraph",
                    "locator_value": locator_value,
                    "source_snippet": "LLM extractor could not produce verified evidence.",
                    "raw_value_text": "",
                    "normalized_value": None,
                    "unit_currency": deal_currency,
                    "extractor_agent_id": "agent_3",
                    "verifier_agent_id": "agent_4",
                    "trace_id": trace_id,
                    "extracted_at": extracted_at,
                },
            }
            rows.append(row)
            continue

        raw_value_text = str(accepted_candidate.get("raw_value_text", ""))
        source_snippet = str(accepted_candidate.get("source_snippet", ""))
        normalization = normalize_value(
            raw_value_text=raw_value_text,
            source_snippet=source_snippet,
            deal_currency=deal_currency,
        )

        base_conf = _safe_float(accepted_candidate.get("confidence", 0.0), default=0.0)
        adjust = 0.0
        if final_verifier is not None:
            try:
                adjust = float(final_verifier.get("confidence_adjustment", 0.0) or 0.0)
            except Exception:
                adjust = 0.0
        confidence = max(0.0, min(1.0, round(base_conf + adjust, 4)))

        blockers: list[str] = []
        if normalization.unresolved_reason:
            blockers.append(normalization.unresolved_reason)

        selected_doc_id = str(accepted_candidate.get("doc_id", "") or fallback_doc_id)
        selected_doc_name = str(accepted_candidate.get("doc_name", "") or fallback_doc_name)
        selected_page = str(accepted_candidate.get("page_or_sheet", "") or fallback_page)
        selected_locator_type = str(accepted_candidate.get("locator_type", "paragraph") or "paragraph")
        selected_locator_value = str(accepted_candidate.get("locator_value", "") or "inferred:missing_locator")

        if not selected_doc_id:
            blockers.append("missing_evidence_location")

        status = classify_status(confidence=confidence, hard_blockers=blockers)
        selected_line = _select_candidate_line(accepted_candidate, candidate_lines)
        reason = _llm_reason_for_row(
            selected_candidate=accepted_candidate,
            selected_line=selected_line,
            candidate_lines=candidate_lines,
            status=status,
            missing_sources=missing_sources,
            normalization_unresolved_reason=normalization.unresolved_reason,
        )

        evidence = {
            "doc_id": selected_doc_id,
            "doc_name": selected_doc_name,
            "page_or_sheet": selected_page,
            "locator_type": selected_locator_type,
            "locator_value": selected_locator_value,
            "source_snippet": source_snippet,
            "raw_value_text": raw_value_text,
            "normalized_value": normalization.normalized_value,
            "unit_currency": normalization.unit_currency,
            "extractor_agent_id": "agent_3",
            "verifier_agent_id": "agent_4",
            "trace_id": trace_id,
            "extracted_at": extracted_at,
        }

        accepted_anchor = {
            "anchor_id": f"{trace_id}:accepted",
            "doc_id": selected_doc_id,
            "doc_name": selected_doc_name,
            "page_or_sheet": selected_page,
            "locator_type": selected_locator_type,
            "locator_value": selected_locator_value,
            "source_snippet": source_snippet,
            "raw_value_text": raw_value_text,
            "normalized_value": normalization.normalized_value,
            "unit_currency": normalization.unit_currency,
            "concept_id": concept_id,
            "concept_label": concept_label,
            "period_id": package_id,
            "trace_id": trace_id,
            "source_role": _source_role_for_doc_type(str(accepted_candidate.get("doc_type", ""))),
            "confidence": confidence,
        }
        source_anchors.append(accepted_anchor)

        rows.append(
            {
                "concept_id": concept_id,
                "label": concept_label,
                "status": status,
                "dictionary_version": "v1.0",
                "raw_value_text": raw_value_text,
                "normalized_value": normalization.normalized_value,
                "current_value": normalization.normalized_value,
                "unit_currency": normalization.unit_currency,
                "confidence": confidence,
                "hard_blockers": sorted(set(blockers)),
                "trace_id": trace_id,
                **reason,
                "source_anchors": unique_trustworthy_anchors(source_anchors, max_items=8),
                "llm_trace": {
                    "model": model_name,
                    "agent_2_classification": file_classifications,
                    "attempts": attempt_logs,
                    "final_reason": "accepted",
                    "max_retries": max_retries,
                },
                "evidence_link": {
                    "doc_id": selected_doc_id,
                    "doc_name": selected_doc_name,
                    "page_or_sheet": selected_page,
                    "locator_type": selected_locator_type,
                    "locator_value": selected_locator_value,
                },
                "evidence": evidence,
            }
        )

    return {
        "package_id": package_id,
        "deal_id": package_manifest["deal_id"],
        "period_end_date": package_manifest["period_end_date"],
        "rows": rows,
    }
