from __future__ import annotations

from dataclasses import dataclass
import hashlib
from pathlib import Path
import re
from typing import Any

from .constants import CONCEPT_LABELS
from .storage import resolve_storage_uri

try:  # pragma: no cover - optional dependency path
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None

try:  # pragma: no cover - optional dependency path
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency path
    PdfReader = None


@dataclass(frozen=True)
class ObligationLine:
    doc_id: str
    doc_name: str
    doc_type: str
    storage_uri: str
    page_or_sheet: str
    locator_type: str
    locator_value: str
    text: str


# Narrow deterministic v1 support: one concept family only.
SUPPORTED_REPORTING_OBLIGATION_CONCEPTS = ("net_income",)

_CONCEPT_PATTERNS = {
    "net_income": (
        re.compile(r"\bnet income\b", re.IGNORECASE),
        re.compile(r"\bnet earnings\b", re.IGNORECASE),
    ),
}

_OBLIGATION_PHRASE_PATTERNS = (
    re.compile(r"\bmust\s+(include|provide|submit|deliver|furnish|report)\b", re.IGNORECASE),
    re.compile(r"\bshall\s+(include|provide|submit|deliver|furnish|report)\b", re.IGNORECASE),
    re.compile(r"\bis required to\s+(include|provide|submit|deliver|furnish|report)\b", re.IGNORECASE),
    re.compile(r"\brequired reporting (package|deliverable)\b", re.IGNORECASE),
)

_REPORTING_CONTEXT_PATTERN = re.compile(
    r"\b(reporting package|reporting deliverable|required reporting|compliance certificate|monthly reporting|quarterly reporting|annual reporting|reporting schedule|financial statement|financial statements)\b",
    re.IGNORECASE,
)

_WEAK_CONTEXT_PATTERN = re.compile(
    r"\b(overview|index|discussion|commentary|summary|illustrative|example|for reference|as applicable)\b",
    re.IGNORECASE,
)


def _safe_text(value: Any, max_len: int = 900) -> str:
    text = str(value or "").strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def detect_reporting_obligation_cadence(text: str) -> str | None:
    lowered = text.lower()
    if "monthly" in lowered:
        return "monthly"
    if "quarterly" in lowered:
        return "quarterly"
    if "annual" in lowered or "annually" in lowered or "yearly" in lowered:
        return "annual"
    return None


def detect_reporting_obligation_source_role(doc_name: str, snippet: str) -> str:
    lowered = f"{doc_name} {snippet}".lower()
    if "reporting" in lowered and "schedule" in lowered:
        return "credit_reporting_schedule"
    if "agreement" in lowered:
        return "credit_agreement_document"
    return "deal_reporting_document"


def reporting_requirement_strength(text: str) -> str:
    lowered = str(text or "").strip()
    has_phrase = any(pattern.search(lowered) for pattern in _OBLIGATION_PHRASE_PATTERNS)
    has_context = _REPORTING_CONTEXT_PATTERN.search(lowered) is not None
    has_weak_context = _WEAK_CONTEXT_PATTERN.search(lowered) is not None
    if has_phrase and has_context:
        if has_weak_context:
            return "ambiguous"
        return "grounded"
    if has_phrase or has_context:
        return "ambiguous"
    return "unsupported"


def build_reporting_obligation_id(
    *,
    deal_id: str,
    doc_id: str,
    locator_type: str,
    locator_value: str,
    required_concept_id: str,
) -> str:
    base = "|".join(
        [
            deal_id.strip().lower(),
            doc_id.strip(),
            locator_type.strip().lower(),
            locator_value.strip(),
            required_concept_id.strip().lower(),
        ]
    )
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]
    return f"obl_{digest}"


def _read_pdf_lines(path: Path, *, doc_id: str, doc_name: str, storage_uri: str) -> list[ObligationLine]:
    if PdfReader is None:
        return []

    reader = PdfReader(str(path))
    rows: list[ObligationLine] = []
    for page_num, page in enumerate(reader.pages[:40], start=1):
        text = page.extract_text() or ""
        if not text.strip():
            continue
        for line_num, line in enumerate(text.splitlines(), start=1):
            clean = _safe_text(line, max_len=700)
            if not clean:
                continue
            rows.append(
                ObligationLine(
                    doc_id=doc_id,
                    doc_name=doc_name,
                    doc_type="PDF",
                    storage_uri=storage_uri,
                    page_or_sheet=f"Page {page_num}",
                    locator_type="paragraph",
                    locator_value=f"p{page_num}:l{line_num}",
                    text=clean,
                )
            )
            if len(rows) >= 1800:
                return rows
    return rows


def _read_xlsx_lines(path: Path, *, doc_id: str, doc_name: str, storage_uri: str) -> list[ObligationLine]:
    if load_workbook is None:
        return []

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    rows: list[ObligationLine] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=1200):
                values: list[str] = []
                first_coordinate = "A1"
                for idx, cell in enumerate(row):
                    if idx == 0:
                        first_coordinate = str(getattr(cell, "coordinate", "A1"))
                    value = cell.value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    values.append(text)
                if not values:
                    continue
                joined = _safe_text(" | ".join(values), max_len=700)
                rows.append(
                    ObligationLine(
                        doc_id=doc_id,
                        doc_name=doc_name,
                        doc_type="XLSX",
                        storage_uri=storage_uri,
                        page_or_sheet=f"Sheet: {sheet.title}",
                        locator_type="cell",
                        locator_value=first_coordinate,
                        text=joined,
                    )
                )
                if len(rows) >= 1800:
                    return rows
    finally:
        workbook.close()
    return rows


def _doc_lines(doc: dict[str, Any]) -> list[ObligationLine]:
    doc_id = str(doc.get("doc_id", "")).strip()
    doc_name = str(doc.get("doc_name", "") or doc.get("filename", "")).strip()
    doc_type = str(doc.get("doc_type", "")).strip().upper()
    storage_uri = str(doc.get("storage_uri", "")).strip()
    if not doc_id or not doc_name or not storage_uri:
        return []

    resolved = resolve_storage_uri(storage_uri)
    if resolved is None:
        return []

    if doc_type == "PDF":
        return _read_pdf_lines(
            resolved,
            doc_id=doc_id,
            doc_name=doc_name,
            storage_uri=storage_uri,
        )
    if doc_type == "XLSX":
        return _read_xlsx_lines(
            resolved,
            doc_id=doc_id,
            doc_name=doc_name,
            storage_uri=storage_uri,
        )
    return []


def _extract_line_obligations(deal_id: str, line: ObligationLine) -> list[dict[str, Any]]:
    text = line.text.strip()
    if not text:
        return []

    obligations: list[dict[str, Any]] = []
    strength = reporting_requirement_strength(text)
    cadence = detect_reporting_obligation_cadence(text)

    for concept_id in SUPPORTED_REPORTING_OBLIGATION_CONCEPTS:
        patterns = _CONCEPT_PATTERNS.get(concept_id, ())
        if not any(pattern.search(text) for pattern in patterns):
            continue
        if strength == "unsupported":
            continue

        obligations.append(
            {
                "obligation_id": build_reporting_obligation_id(
                    deal_id=deal_id,
                    doc_id=line.doc_id,
                    locator_type=line.locator_type,
                    locator_value=line.locator_value,
                    required_concept_id=concept_id,
                ),
                "deal_id": deal_id,
                "doc_id": line.doc_id,
                "doc_name": line.doc_name,
                "doc_type": line.doc_type,
                "storage_uri": line.storage_uri,
                "locator_type": line.locator_type,
                "locator_value": line.locator_value,
                "page_or_sheet": line.page_or_sheet,
                "source_snippet": text,
                "obligation_type": "reporting_requirement",
                "required_concept_id": concept_id,
                "required_concept_label": CONCEPT_LABELS.get(concept_id, concept_id),
                "cadence": cadence,
                "source_role": detect_reporting_obligation_source_role(line.doc_name, text),
                "grounding_state": strength,
            }
        )
    return obligations


def extract_reporting_obligations(
    *,
    deal_id: str,
    docs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for doc in docs:
        for line in _doc_lines(doc):
            candidates.extend(_extract_line_obligations(deal_id=deal_id, line=line))

    # Deterministic dedupe: prefer grounded over ambiguous for the same locator/concept.
    deduped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for item in candidates:
        key = (
            str(item.get("doc_id", "")).strip(),
            str(item.get("locator_type", "")).strip().lower(),
            str(item.get("locator_value", "")).strip(),
            str(item.get("required_concept_id", "")).strip().lower(),
        )
        current = deduped.get(key)
        if current is None:
            deduped[key] = item
            continue
        current_state = str(current.get("grounding_state", "unsupported"))
        next_state = str(item.get("grounding_state", "unsupported"))
        if current_state != "grounded" and next_state == "grounded":
            deduped[key] = item

    return sorted(
        deduped.values(),
        key=lambda item: (
            0 if str(item.get("grounding_state", "")) == "grounded" else 1,
            str(item.get("required_concept_id", "")),
            str(item.get("doc_id", "")),
            str(item.get("locator_value", "")),
        ),
    )


def collect_reporting_obligation_lines(docs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lines: list[dict[str, Any]] = []
    for doc in docs:
        for line in _doc_lines(doc):
            lines.append(
                {
                    "doc_id": line.doc_id,
                    "doc_name": line.doc_name,
                    "doc_type": line.doc_type,
                    "storage_uri": line.storage_uri,
                    "page_or_sheet": line.page_or_sheet,
                    "locator_type": line.locator_type,
                    "locator_value": line.locator_value,
                    "text": line.text,
                }
            )
    return lines
