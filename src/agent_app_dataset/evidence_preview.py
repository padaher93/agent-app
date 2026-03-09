from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from .storage import resolve_storage_uri as resolve_storage_uri_path

try:  # pragma: no cover - optional dependency path
    from openpyxl import load_workbook
    from openpyxl.utils.cell import get_column_letter
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None
    get_column_letter = None

try:  # pragma: no cover - optional dependency path
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency path
    PdfReader = None


def resolve_storage_uri(uri: str) -> Path | None:
    return resolve_storage_uri_path(uri)


def _column_index(col: str) -> int:
    value = 0
    for char in col.upper():
        value = value * 26 + (ord(char) - 64)
    return max(value, 1)


def _parse_cell_locator(locator: str) -> tuple[int, int] | None:
    m = re.match(r"^([A-Za-z]+)([0-9]+)$", str(locator).strip())
    if not m:
        return None
    col = _column_index(m.group(1))
    row = int(m.group(2))
    return row, col


def _xlsx_grid_preview(path: Path, sheet_hint: str, locator: str) -> dict[str, Any]:
    if load_workbook is None:
        return {"kind": "none", "reason": "openpyxl_unavailable"}

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        sheet_name = sheet_hint.replace("Sheet: ", "").strip() if sheet_hint else ""
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.worksheets[0]

        target = _parse_cell_locator(locator) or (1, 1)
        row_idx, col_idx = target

        max_row = max(1, int(getattr(sheet, "max_row", 1) or 1))
        max_col = max(1, int(getattr(sheet, "max_column", 1) or 1))

        row_cap = 300
        col_cap = 80

        start_row = 1
        end_row = min(max_row, row_cap)
        if row_idx > end_row:
            start_row = max(1, row_idx - row_cap + 1)
            end_row = min(max_row, start_row + row_cap - 1)

        start_col = 1
        end_col = min(max_col, col_cap)
        if col_idx > end_col:
            start_col = max(1, col_idx - col_cap + 1)
            end_col = min(max_col, start_col + col_cap - 1)

        rows: list[list[dict[str, Any]]] = []
        for row_offset, row in enumerate(
            sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
            ),
            start=0,
        ):
            line: list[dict[str, Any]] = []
            for col_offset, cell in enumerate(row, start=0):
                abs_row = start_row + row_offset
                abs_col = start_col + col_offset
                if get_column_letter:
                    coordinate = f"{get_column_letter(abs_col)}{abs_row}"
                else:
                    coordinate = f"R{abs_row}C{abs_col}"
                value = getattr(cell, "value", None)
                line.append(
                    {
                        "coordinate": coordinate,
                        "value": value,
                        "highlight": abs_row == row_idx and abs_col == col_idx,
                    }
                )
            rows.append(line)

        return {
            "kind": "xlsx_sheet",
            "sheet": sheet.title,
            "target": locator,
            "total_rows": max_row,
            "total_cols": max_col,
            "viewport_start_row": start_row,
            "viewport_end_row": end_row,
            "viewport_start_col": start_col,
            "viewport_end_col": end_col,
            "truncated": max_row > row_cap or max_col > col_cap,
            "rows": rows,
        }
    finally:
        workbook.close()


def _pdf_text_preview(path: Path, locator: str) -> dict[str, Any]:
    if PdfReader is None:
        return {"kind": "none", "reason": "pypdf_unavailable"}

    reader = PdfReader(str(path))
    page_no = 1
    m = re.match(r"p([0-9]+):", locator)
    if m:
        page_no = max(1, int(m.group(1)))

    page_no = min(page_no, len(reader.pages))
    text = (reader.pages[page_no - 1].extract_text() or "").strip()
    return {
        "kind": "pdf_text",
        "page": page_no,
        "text": text[:4000],
    }


def build_evidence_preview(
    trace_row: dict[str, Any],
    package_manifest: dict[str, Any],
) -> dict[str, Any]:
    evidence = trace_row.get("evidence", {})
    doc_id = str(evidence.get("doc_id", ""))
    locator_type = str(evidence.get("locator_type", "paragraph"))
    locator_value = str(evidence.get("locator_value", ""))

    file_meta = None
    for file in package_manifest.get("files", []):
        if str(file.get("file_id", "")) == doc_id:
            file_meta = file
            break

    if file_meta is None:
        return {
            "available": False,
            "reason": "document_not_in_manifest",
            "doc_id": doc_id,
            "locator_type": locator_type,
            "locator_value": locator_value,
        }

    return build_document_locator_preview(
        doc_id=doc_id,
        doc_type=str(file_meta.get("doc_type", "")).upper(),
        filename=str(file_meta.get("filename", "")),
        storage_uri=str(file_meta.get("storage_uri", "")),
        locator_type=locator_type,
        locator_value=locator_value,
        page_or_sheet=str(evidence.get("page_or_sheet", "")),
        source_snippet=str(evidence.get("source_snippet", "")),
    )


def build_document_locator_preview(
    *,
    doc_id: str,
    doc_type: str,
    filename: str,
    storage_uri: str,
    locator_type: str,
    locator_value: str,
    page_or_sheet: str = "",
    source_snippet: str = "",
) -> dict[str, Any]:
    resolved_path = resolve_storage_uri(str(storage_uri))
    if resolved_path is None:
        return {
            "available": False,
            "reason": "document_unavailable",
            "doc_id": str(doc_id),
            "doc_type": str(doc_type).upper(),
            "filename": str(filename),
            "storage_uri": str(storage_uri),
            "locator_type": str(locator_type),
            "locator_value": str(locator_value),
            "page_or_sheet": str(page_or_sheet),
            "source_snippet": str(source_snippet),
            "preview": {"kind": "none", "reason": "document_unavailable"},
        }

    normalized_doc_type = str(doc_type).upper()
    if normalized_doc_type == "XLSX":
        preview = _xlsx_grid_preview(
            path=resolved_path,
            sheet_hint=str(page_or_sheet),
            locator=str(locator_value),
        )
    elif normalized_doc_type == "PDF":
        preview = _pdf_text_preview(path=resolved_path, locator=str(locator_value))
    else:
        preview = {"kind": "none", "reason": "unsupported_doc_type"}

    return {
        "available": True,
        "doc_id": str(doc_id),
        "doc_type": normalized_doc_type,
        "filename": str(filename),
        "storage_uri": str(storage_uri),
        "locator_type": str(locator_type),
        "locator_value": str(locator_value),
        "page_or_sheet": str(page_or_sheet),
        "source_snippet": str(source_snippet),
        "preview": preview,
    }
