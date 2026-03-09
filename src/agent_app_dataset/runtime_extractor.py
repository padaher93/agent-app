from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any

from .constants import CONCEPT_DEFINITIONS, CONCEPT_LABELS, STARTER_CONCEPT_IDS
from .normalization import normalize_value
from .policy import classify_status
from .source_grounding import unique_trustworthy_anchors
from .storage import resolve_storage_uri

try:  # pragma: no cover - optional dependency path
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None

try:  # pragma: no cover - optional dependency path
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency path
    PdfReader = None


@dataclass(frozen=True)
class ConceptCandidate:
    concept_id: str
    raw_value_text: str
    source_snippet: str
    doc_id: str
    doc_name: str
    page_or_sheet: str
    locator_type: str
    locator_value: str
    confidence: float
    match_basis: str
    source_modality: str


def _source_role_for_locator(locator_type: str) -> str:
    if str(locator_type).strip().lower() == "cell":
        return "worksheet_value"
    return "submitted_source_line"


def _build_source_anchors(
    *,
    concept_id: str,
    concept_label: str,
    package_id: str,
    trace_id: str,
    candidates: list[ConceptCandidate],
    deal_currency: str,
) -> list[dict[str, Any]]:
    anchors: list[dict[str, Any]] = []
    for index, candidate in enumerate(candidates, start=1):
        normalized = normalize_value(
            raw_value_text=candidate.raw_value_text,
            source_snippet=candidate.source_snippet,
            deal_currency=deal_currency,
        )
        anchors.append(
            {
                "anchor_id": f"{trace_id}:cand:{index}",
                "doc_id": candidate.doc_id,
                "doc_name": candidate.doc_name,
                "page_or_sheet": candidate.page_or_sheet,
                "locator_type": candidate.locator_type,
                "locator_value": candidate.locator_value,
                "source_snippet": candidate.source_snippet,
                "raw_value_text": candidate.raw_value_text,
                "normalized_value": normalized.normalized_value,
                "unit_currency": normalized.unit_currency,
                "concept_id": concept_id,
                "concept_label": concept_label,
                "period_id": package_id,
                "trace_id": trace_id,
                "source_role": _source_role_for_locator(candidate.locator_type),
                "confidence": candidate.confidence,
            }
        )

    return unique_trustworthy_anchors(anchors, max_items=8)


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _extract_numeric_text(value: Any) -> str | None:
    if isinstance(value, (int, float)):
        return str(value)
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    m = re.search(r"-?[0-9][0-9,]*(?:\.[0-9]+)?", text)
    if not m:
        return None
    return m.group(0)


def _keyword_score(line: str, keywords: tuple[str, ...]) -> float:
    lowered = line.lower()
    if not lowered:
        return 0.0

    score = 0.0
    for keyword in keywords:
        kw = keyword.lower()
        if kw == lowered.strip():
            score = max(score, 1.0)
        elif kw in lowered:
            score = max(score, 0.92)
    return score


def _candidate_from_text(
    concept_id: str,
    line: str,
    doc_id: str,
    doc_name: str,
    page_or_sheet: str,
    locator_type: str,
    locator_value: str,
    source_modality: str,
) -> ConceptCandidate | None:
    keywords = tuple(CONCEPT_DEFINITIONS[concept_id]["keywords"])
    keyword_score = _keyword_score(line, keywords)
    if keyword_score <= 0:
        return None

    numeric_text = _extract_numeric_text(line)
    if numeric_text:
        confidence = min(0.97, keyword_score + 0.02)
        raw_value_text = numeric_text
    else:
        confidence = 0.81
        raw_value_text = ""
    match_basis = "exact_label_match" if keyword_score >= 0.99 else "label_variant_match"

    return ConceptCandidate(
        concept_id=concept_id,
        raw_value_text=raw_value_text,
        source_snippet=line[:500],
        doc_id=doc_id,
        doc_name=doc_name,
        page_or_sheet=page_or_sheet,
        locator_type=locator_type,
        locator_value=locator_value,
        confidence=round(confidence, 4),
        match_basis=match_basis,
        source_modality=source_modality,
    )


def _reason_payload(
    *,
    code: str | None,
    label: str | None,
    detail: str | None,
    uncertainty_source: str | None,
    match_basis: str | None,
    source_modality: str | None,
    candidate_count: int,
    expected_section_state: str | None = None,
) -> dict[str, Any]:
    return {
        "extraction_reason_code": code,
        "extraction_reason_label": label,
        "extraction_reason_detail": detail,
        "uncertainty_source": uncertainty_source,
        "match_basis": match_basis,
        "source_modality": source_modality,
        "candidate_count": int(max(candidate_count, 0)),
        "expected_section_state": expected_section_state,
    }


def _runtime_reason_for_row(
    *,
    selected: ConceptCandidate | None,
    candidate_count: int,
    status: str,
    missing_sources: bool,
    normalization_unresolved_reason: str | None,
    selected_locator_value: str,
) -> dict[str, Any]:
    reason_code: str | None = None
    reason_label: str | None = None
    reason_detail: str | None = None
    uncertainty_source: str | None = None
    match_basis = selected.match_basis if selected is not None else None
    source_modality = selected.source_modality if selected is not None else "none"
    expected_section_state: str | None = None

    if selected is None:
        if missing_sources:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "Source files were unavailable during extraction."
            uncertainty_source = "package_extraction"
            expected_section_state = "source_document_unavailable"
        else:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "No exact source anchor was found in the current package."
            uncertainty_source = "package_extraction"
            expected_section_state = "exact_anchor_not_found"
        return _reason_payload(
            code=reason_code,
            label=reason_label,
            detail=reason_detail,
            uncertainty_source=uncertainty_source,
            match_basis=match_basis,
            source_modality=source_modality,
            candidate_count=candidate_count,
            expected_section_state=expected_section_state,
        )

    if status in {"candidate_flagged", "unresolved"}:
        if not selected_locator_value or selected_locator_value.startswith("inferred:"):
            reason_code = "exact_row_header_missing"
            reason_label = "Exact row header missing"
            reason_detail = "Candidate value was found without a precise structured row locator."
            uncertainty_source = "package_extraction"
            expected_section_state = "structured_locator_missing"
        elif source_modality == "pdf_text":
            reason_code = "candidate_from_pdf_text_only"
            reason_label = "Extracted from PDF text only"
            reason_detail = "Candidate was found in PDF text, not in a structured table row."
            uncertainty_source = "package_extraction"
        elif source_modality == "narrative_text":
            reason_code = "candidate_from_narrative_text"
            reason_label = "Candidate extracted from narrative text"
            reason_detail = "Candidate came from narrative text rather than a structured table row."
            uncertainty_source = "package_extraction"
        elif candidate_count >= 2:
            reason_code = "multiple_matching_rows"
            reason_label = "Multiple matching rows"
            reason_detail = "More than one matching anchor was found for this concept."
            uncertainty_source = "package_extraction"
        elif match_basis == "label_variant_match":
            reason_code = "label_variant_match"
            reason_label = "Label variant match"
            reason_detail = "Match used a label variant instead of an exact header match."
            uncertainty_source = "package_extraction"
        elif normalization_unresolved_reason:
            reason_code = "current_package_missing_exact_support"
            reason_label = "Current package missing exact support"
            reason_detail = "Extracted candidate could not be normalized into a trustworthy numeric value."
            uncertainty_source = "package_extraction"

    return _reason_payload(
        code=reason_code,
        label=reason_label,
        detail=reason_detail,
        uncertainty_source=uncertainty_source,
        match_basis=match_basis,
        source_modality=source_modality,
        candidate_count=candidate_count,
        expected_section_state=expected_section_state,
    )


def _read_xlsx_candidates(path: Path, file_meta: dict[str, Any]) -> list[ConceptCandidate]:
    if load_workbook is None:
        return []

    candidates: list[ConceptCandidate] = []
    workbook = load_workbook(filename=path, data_only=True, read_only=True)

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=500):
                row_values: list[str] = []
                numeric_by_col: dict[int, str] = {}
                for idx, cell in enumerate(row):
                    value = cell.value
                    if value is None:
                        continue
                    if isinstance(value, str):
                        row_values.append(value)
                    elif isinstance(value, (int, float)):
                        row_values.append(str(value))
                    numeric_text = _extract_numeric_text(value)
                    if numeric_text:
                        numeric_by_col[idx] = numeric_text

                if not row_values:
                    continue

                combined = " | ".join(row_values)
                for concept_id in STARTER_CONCEPT_IDS:
                    first_cell = row[0] if row else None
                    locator_value = getattr(first_cell, "coordinate", "") if first_cell is not None else ""
                    base = _candidate_from_text(
                        concept_id=concept_id,
                        line=combined,
                        doc_id=str(file_meta.get("file_id", "")),
                        doc_name=str(file_meta.get("filename", path.name)),
                        page_or_sheet=f"Sheet: {sheet.title}",
                        locator_type="cell",
                        locator_value=locator_value,
                        source_modality="table_cell",
                    )
                    if base is None:
                        continue

                    if not base.raw_value_text and numeric_by_col:
                        first_numeric = next(iter(numeric_by_col.values()))
                        base = ConceptCandidate(
                            concept_id=base.concept_id,
                            raw_value_text=first_numeric,
                            source_snippet=base.source_snippet,
                            doc_id=base.doc_id,
                            doc_name=base.doc_name,
                            page_or_sheet=base.page_or_sheet,
                            locator_type=base.locator_type,
                            locator_value=base.locator_value,
                            confidence=0.85,
                            match_basis=base.match_basis,
                            source_modality=base.source_modality,
                        )
                    candidates.append(base)
    finally:
        workbook.close()

    return candidates


def _read_pdf_candidates(path: Path, file_meta: dict[str, Any]) -> list[ConceptCandidate]:
    if PdfReader is None:
        return []

    candidates: list[ConceptCandidate] = []
    reader = PdfReader(str(path))

    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if not text.strip():
            continue

        for line_num, line in enumerate(text.splitlines(), start=1):
            clean = line.strip()
            if not clean:
                continue

            for concept_id in STARTER_CONCEPT_IDS:
                candidate = _candidate_from_text(
                    concept_id=concept_id,
                    line=clean,
                    doc_id=str(file_meta.get("file_id", "")),
                    doc_name=str(file_meta.get("filename", path.name)),
                    page_or_sheet=f"Page {page_num}",
                    locator_type="paragraph",
                    locator_value=f"p{page_num}:l{line_num}",
                    source_modality="pdf_text",
                )
                if candidate is not None:
                    candidates.append(candidate)

    return candidates


def _collect_candidates(package_manifest: dict[str, Any]) -> tuple[dict[str, list[ConceptCandidate]], bool]:
    by_concept: dict[str, list[ConceptCandidate]] = {concept_id: [] for concept_id in STARTER_CONCEPT_IDS}
    missing_sources = False

    for file_meta in package_manifest.get("files", []):
        doc_type = str(file_meta.get("doc_type", "")).upper()
        path = resolve_storage_uri(str(file_meta.get("storage_uri", "")))
        if path is None:
            missing_sources = True
            continue

        if doc_type == "XLSX":
            extracted = _read_xlsx_candidates(path=path, file_meta=file_meta)
        elif doc_type == "PDF":
            extracted = _read_pdf_candidates(path=path, file_meta=file_meta)
        else:
            extracted = []

        for item in extracted:
            by_concept[item.concept_id].append(item)

    return by_concept, missing_sources


def _best_candidate(candidates: list[ConceptCandidate]) -> ConceptCandidate | None:
    if not candidates:
        return None
    ranked = sorted(
        candidates,
        key=lambda item: (item.confidence, bool(item.raw_value_text), item.doc_id),
        reverse=True,
    )
    return ranked[0]


def runtime_extract_package_predictions(
    package_manifest: dict[str, Any],
    deal_currency: str = "USD",
) -> dict[str, Any]:
    package_id = str(package_manifest["package_id"])
    files = package_manifest.get("files", [])
    fallback_file = files[0] if files else {}
    fallback_doc_id = str(fallback_file.get("file_id", ""))
    fallback_doc_name = str(fallback_file.get("filename", ""))
    fallback_doc_type = str(fallback_file.get("doc_type", "")).upper()
    fallback_page = "Page 1"
    if fallback_doc_type == "XLSX":
        fallback_page = "Sheet: unknown"

    concept_candidates, missing_sources = _collect_candidates(package_manifest)

    rows: list[dict[str, Any]] = []
    extracted_at = _now_utc()

    for concept_id in STARTER_CONCEPT_IDS:
        trace_id = f"tr_{package_id}_{concept_id}"
        candidates = concept_candidates.get(concept_id, [])
        source_anchors = _build_source_anchors(
            concept_id=concept_id,
            concept_label=CONCEPT_LABELS[concept_id],
            package_id=package_id,
            trace_id=trace_id,
            candidates=candidates,
            deal_currency=deal_currency,
        )
        selected = _best_candidate(candidates)

        if selected is None:
            blockers = ["missing_numeric_value", "no_reliable_candidate"]
            if missing_sources:
                blockers.append("missing_source_document")

            locator_value = "unresolved:not_found"
            if not fallback_doc_id:
                locator_value = "unresolved:missing_package_file"

            reason = _runtime_reason_for_row(
                selected=None,
                candidate_count=len(candidates),
                status="unresolved",
                missing_sources=missing_sources,
                normalization_unresolved_reason=None,
                selected_locator_value=locator_value,
            )
            row = {
                "concept_id": concept_id,
                "label": CONCEPT_LABELS[concept_id],
                "status": "unresolved",
                "dictionary_version": "v1.0",
                "raw_value_text": "",
                "normalized_value": None,
                "current_value": None,
                "unit_currency": deal_currency,
                "confidence": 0.0,
                "hard_blockers": blockers,
                "trace_id": trace_id,
                **reason,
                "source_anchors": source_anchors,
                "evidence_link": {
                    "doc_id": fallback_doc_id,
                    "doc_name": fallback_doc_name,
                    "page_or_sheet": fallback_page if fallback_doc_id else "Package Context",
                    "locator_type": "paragraph",
                    "locator_value": locator_value,
                },
                "evidence": {
                    "doc_id": fallback_doc_id,
                    "doc_name": fallback_doc_name,
                    "page_or_sheet": fallback_page if fallback_doc_id else "Package Context",
                    "locator_type": "paragraph",
                    "locator_value": locator_value,
                    "source_snippet": "No reliable candidate found. Row is anchored to package context for auditability.",
                    "raw_value_text": "",
                    "normalized_value": None,
                    "unit_currency": deal_currency,
                    "extractor_agent_id": "agent_3",
                    "verifier_agent_id": "agent_4",
                    "trace_id": trace_id,
                    "extracted_at": extracted_at,
                },
            }
            rows.append(row)
            continue

        normalization = normalize_value(
            raw_value_text=selected.raw_value_text,
            source_snippet=selected.source_snippet,
            deal_currency=deal_currency,
        )

        selected_doc_id = selected.doc_id or fallback_doc_id
        selected_doc_name = selected.doc_name or fallback_doc_name
        selected_page = selected.page_or_sheet or (fallback_page if selected_doc_id else "Package Context")
        selected_locator_type = selected.locator_type or "paragraph"
        selected_locator_value = selected.locator_value or "inferred:missing_locator"

        blockers: list[str] = []
        if normalization.unresolved_reason:
            blockers.append(normalization.unresolved_reason)
        if not selected.doc_id or not selected.locator_value:
            blockers.append("weak_evidence_locator")

        status = classify_status(confidence=selected.confidence, hard_blockers=blockers)
        reason = _runtime_reason_for_row(
            selected=selected,
            candidate_count=len(candidates),
            status=status,
            missing_sources=missing_sources,
            normalization_unresolved_reason=normalization.unresolved_reason,
            selected_locator_value=selected_locator_value,
        )
        normalized_value = normalization.normalized_value
        evidence = {
            "doc_id": selected_doc_id,
            "doc_name": selected_doc_name,
            "page_or_sheet": selected_page,
            "locator_type": selected_locator_type,
            "locator_value": selected_locator_value,
            "source_snippet": selected.source_snippet,
            "raw_value_text": selected.raw_value_text,
            "normalized_value": normalized_value,
            "unit_currency": normalization.unit_currency,
            "extractor_agent_id": "agent_3",
            "verifier_agent_id": "agent_4",
            "trace_id": trace_id,
            "extracted_at": extracted_at,
        }

        rows.append(
            {
                "concept_id": concept_id,
                "label": CONCEPT_LABELS[concept_id],
                "status": status,
                "dictionary_version": "v1.0",
                "raw_value_text": selected.raw_value_text,
                "normalized_value": normalized_value,
                "current_value": normalized_value,
                "unit_currency": normalization.unit_currency,
                "confidence": round(float(selected.confidence), 4),
                "hard_blockers": sorted(set(blockers)),
                "trace_id": trace_id,
                **reason,
                "source_anchors": source_anchors,
                "evidence_link": {
                    "doc_id": selected_doc_id,
                    "doc_name": selected_doc_name,
                    "page_or_sheet": selected_page,
                    "locator_type": selected_locator_type,
                    "locator_value": selected_locator_value,
                },
                "evidence": evidence,
            }
        )

    return {
        "package_id": package_id,
        "deal_id": package_manifest["deal_id"],
        "period_end_date": package_manifest["period_end_date"],
        "rows": rows,
    }
