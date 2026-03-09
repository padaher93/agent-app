from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from agent_app_dataset.internal_api import create_app


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Revenue"
    ws["B2"] = 1000
    ws["A3"] = "EBITDA"
    ws["B3"] = 450
    wb.save(path)


def _build_requirement_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws["A1"] = "Borrower shall provide Net Income with each quarterly reporting package."
    wb.save(path)


def _build_value_workbook(path: Path, revenue_value: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Revenue (Total)"
    ws["B2"] = revenue_value
    wb.save(path)


def test_evidence_endpoint_exposes_download_url_and_file_stream(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_workbook(workbook_path)

    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    ingest_payload = {
        "workspace_id": "ws_default",
        "sender_email": "ops@borrower.com",
        "source_email_id": "email_view_001",
        "deal_id": "deal_view",
        "period_end_date": "2026-01-31",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_view_01",
                "source_id": "src_view_01",
                "doc_type": "XLSX",
                "filename": "borrower_update.xlsx",
                "storage_uri": str(workbook_path),
                "checksum": "checksum-view-01",
                "pages_or_sheets": 1,
            }
        ],
    }

    ingest = client.post("/internal/v1/packages:ingest", json=ingest_payload)
    assert ingest.status_code == 200
    package_id = ingest.json()["package_id"]

    process = client.post(
        f"/internal/v1/packages/{package_id}:process",
        json={"async_mode": False, "max_retries": 2, "extraction_mode": "runtime"},
    )
    assert process.status_code == 200

    delta = client.get(f"/internal/v1/deals/deal_view/periods/{package_id}/delta")
    assert delta.status_code == 200
    trace_id = delta.json()["rows"][0]["trace_id"]

    evidence = client.get(f"/internal/v1/traces/{trace_id}/evidence")
    assert evidence.status_code == 200
    preview = evidence.json()["evidence_preview"]
    assert preview["doc_id"] == "file_view_01"
    assert preview["download_url"] == f"/internal/v1/packages/{package_id}/files/file_view_01:download"

    download = client.get(preview["download_url"])
    assert download.status_code == 200
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    blocked = client.get(preview["download_url"], headers={"X-Workspace-Id": "ws_other"})
    assert blocked.status_code == 404


def test_package_file_evidence_preview_uses_file_specific_anchor_locator(tmp_path: Path) -> None:
    source_a = tmp_path / "source_a.xlsx"
    source_b = tmp_path / "source_b.xlsx"
    _build_value_workbook(source_a, 12_450_000)
    _build_value_workbook(source_b, 12_150_000)

    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    ingest_payload = {
        "workspace_id": "ws_default",
        "sender_email": "ops@borrower.com",
        "source_email_id": "email_preview_conflict_001",
        "deal_id": "deal_preview_conflict",
        "period_end_date": "2026-01-31",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_source_a",
                "source_id": "src_source_a",
                "doc_type": "XLSX",
                "filename": source_a.name,
                "storage_uri": str(source_a),
                "checksum": "checksum_source_a",
                "pages_or_sheets": 1,
            },
            {
                "file_id": "file_source_b",
                "source_id": "src_source_b",
                "doc_type": "XLSX",
                "filename": source_b.name,
                "storage_uri": str(source_b),
                "checksum": "checksum_source_b",
                "pages_or_sheets": 1,
            },
        ],
    }
    ingest = client.post("/internal/v1/packages:ingest", json=ingest_payload)
    assert ingest.status_code == 200
    package_id = ingest.json()["package_id"]

    def highlight_value(payload: dict) -> str:
        rows = payload["evidence_preview"]["preview"]["rows"]
        for row in rows:
            for cell in row:
                if cell.get("highlight"):
                    return str(cell.get("value"))
        return ""

    preview_a = client.get(
        f"/internal/v1/packages/{package_id}/files/file_source_a/evidence-preview",
        params={
            "locator_type": "cell",
            "locator_value": "B2",
            "page_or_sheet": "Sheet: Coverage",
        },
    )
    assert preview_a.status_code == 200
    payload_a = preview_a.json()
    assert payload_a["evidence_preview"]["doc_id"] == "file_source_a"
    assert payload_a["evidence_preview"]["preview"]["kind"] == "xlsx_sheet"
    assert highlight_value(payload_a) == "12450000"

    preview_b = client.get(
        f"/internal/v1/packages/{package_id}/files/file_source_b/evidence-preview",
        params={
            "locator_type": "cell",
            "locator_value": "B2",
            "page_or_sheet": "Sheet: Coverage",
        },
    )
    assert preview_b.status_code == 200
    payload_b = preview_b.json()
    assert payload_b["evidence_preview"]["doc_id"] == "file_source_b"
    assert payload_b["evidence_preview"]["preview"]["kind"] == "xlsx_sheet"
    assert highlight_value(payload_b) == "12150000"


def test_reporting_requirement_preview_endpoint_returns_anchor_preview(tmp_path: Path) -> None:
    requirement_path = tmp_path / "reporting_requirements.xlsx"
    _build_requirement_workbook(requirement_path)

    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    create_deal = client.post(
        "/internal/v1/deals",
        json={"display_name": "Deal Requirement Preview", "deal_id": "deal_req_preview"},
    )
    assert create_deal.status_code == 200

    ingest = client.post(
        "/internal/v1/deals/deal_req_preview/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_preview_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_preview_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "off",
        },
    )
    assert ingest.status_code == 200
    assert ingest.json()["grounded_extracted"] >= 1

    obligations = client.get("/internal/v1/deals/deal_req_preview/reporting-obligations")
    assert obligations.status_code == 200
    grounded = [
        row
        for row in obligations.json()["obligations"]
        if row["grounding_state"] == "grounded" and row["required_concept_id"] == "net_income"
    ]
    assert grounded
    obligation_id = grounded[0]["obligation_id"]

    preview = client.get(
        f"/internal/v1/deals/deal_req_preview/reporting-obligations/{obligation_id}/preview"
    )
    assert preview.status_code == 200
    payload = preview.json()
    assert payload["obligation_id"] == obligation_id
    assert payload["evidence_preview"]["doc_id"] == "req_doc_preview_01"
    assert payload["evidence_preview"]["locator_type"]
    assert payload["evidence_preview"]["locator_value"]
    assert payload["evidence_preview"]["source_snippet"]
    assert payload["evidence_preview"]["download_url"].endswith("/document:download")
