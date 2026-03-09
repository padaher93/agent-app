from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from agent_app_dataset.constants import STARTER_CONCEPT_IDS
from agent_app_dataset.internal_processing import process_package_manifest
from agent_app_dataset.llm_runtime import run_llm_multi_agent_extraction


class _FakeLLMClient:
    def run_json(self, *, agent_id: str, system_prompt: str, user_payload: dict):
        if agent_id == "agent_2":
            files = user_payload.get("files", [])
            return {
                "files": [
                    {
                        "doc_id": item.get("doc_id", ""),
                        "classification": "financial_statement",
                        "confidence": 0.95,
                        "rationale": "contains financial rows",
                    }
                    for item in files
                ]
            }

        if agent_id == "agent_3":
            concept_id = user_payload["concept"]["concept_id"]
            if concept_id == "ebitda_reported":
                return {
                    "candidate": {
                        "found": True,
                        "raw_value_text": "1500000",
                        "doc_id": "file_01",
                        "doc_name": "borrower_update.xlsx",
                        "page_or_sheet": "Sheet: Financials",
                        "locator_type": "cell",
                        "locator_value": "A2",
                        "source_snippet": "EBITDA Reported 1,500,000",
                        "confidence": 0.93,
                        "reason": "exact keyword match",
                    }
                }
            return {
                "candidate": {
                    "found": False,
                    "reason": "not_found_in_documents",
                }
            }

        if agent_id == "agent_4":
            concept_id = user_payload["concept"]["concept_id"]
            if concept_id == "ebitda_reported":
                return {
                    "decision": "accept",
                    "confidence_adjustment": 0.03,
                    "objections": [],
                    "reason": "evidence aligns with concept",
                }
            return {
                "decision": "reject",
                "confidence_adjustment": -0.2,
                "objections": ["wrong_context"],
                "reason": "rejected",
            }

        raise AssertionError(f"unexpected agent id {agent_id}")


class _StubLLMClient:
    def run_json(self, *, agent_id: str, system_prompt: str, user_payload: dict):
        if agent_id == "agent_2":
            return {"files": []}
        if agent_id == "agent_3":
            return {
                "candidate": {
                    "found": False,
                    "reason": "no_docs",
                }
            }
        if agent_id == "agent_4":
            return {
                "decision": "reject",
                "confidence_adjustment": 0,
                "objections": ["no_docs"],
                "reason": "no_docs",
            }
        raise AssertionError(agent_id)


class _FakeLLMUnanchoredClient:
    def run_json(self, *, agent_id: str, system_prompt: str, user_payload: dict):
        if agent_id == "agent_2":
            files = user_payload.get("files", [])
            return {
                "files": [
                    {
                        "doc_id": item.get("doc_id", ""),
                        "classification": "financial_statement",
                        "confidence": 0.91,
                        "rationale": "contains financial rows",
                    }
                    for item in files
                ]
            }
        if agent_id == "agent_3":
            concept_id = user_payload["concept"]["concept_id"]
            if concept_id == "ebitda_reported":
                return {
                    "candidate": {
                        "found": True,
                        "raw_value_text": "1500000",
                        "doc_id": "file_pdf",
                        "doc_name": "borrower_update.pdf",
                        "page_or_sheet": "Page 2",
                        "locator_type": "paragraph",
                        "locator_value": "",
                        "source_snippet": "EBITDA Reported 1,500,000 appears in management discussion.",
                        "confidence": 0.82,
                        "reason": "candidate extracted from narrative text",
                    }
                }
            return {"candidate": {"found": False, "reason": "not_found_in_documents"}}
        if agent_id == "agent_4":
            return {
                "decision": "accept",
                "confidence_adjustment": 0.0,
                "objections": [],
                "reason": "acceptable candidate",
            }
        raise AssertionError(agent_id)


class _FakeLLMPdfCandidateClient:
    def run_json(self, *, agent_id: str, system_prompt: str, user_payload: dict):
        if agent_id == "agent_2":
            files = user_payload.get("files", [])
            return {
                "files": [
                    {
                        "doc_id": item.get("doc_id", ""),
                        "classification": "financial_statement",
                        "confidence": 0.9,
                        "rationale": "contains metrics in text",
                    }
                    for item in files
                ]
            }
        if agent_id == "agent_3":
            concept_id = user_payload["concept"]["concept_id"]
            if concept_id == "ebitda_reported":
                return {
                    "candidate": {
                        "found": True,
                        "raw_value_text": "1500000",
                        "doc_id": "file_pdf",
                        "doc_name": "borrower_update.pdf",
                        "page_or_sheet": "Page 3",
                        "locator_type": "paragraph",
                        "locator_value": "p3:l14",
                        "source_snippet": "EBITDA Reported 1,500,000 from narrative section.",
                        "confidence": 0.83,
                        "reason": "extracted from narrative paragraph",
                    }
                }
            return {"candidate": {"found": False, "reason": "not_found_in_documents"}}
        if agent_id == "agent_4":
            return {
                "decision": "accept",
                "confidence_adjustment": 0.0,
                "objections": [],
                "reason": "accepted",
            }
        raise AssertionError(agent_id)


def _build_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "EBITDA Reported"
    ws["B2"] = 1500000
    ws["A3"] = "Total Debt"
    ws["B3"] = 7200000
    wb.save(path)


def _manifest(path: Path) -> dict:
    return {
        "package_id": "pkg_llm_001",
        "workspace_id": "ws_default",
        "deal_id": "deal_llm",
        "period_end_date": "2026-01-31",
        "source_email_id": "email_llm_001",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_01",
                "source_id": "src_01",
                "doc_type": "XLSX",
                "filename": "borrower_update.xlsx",
                "storage_uri": str(path),
                "checksum": "checksum_01",
                "pages_or_sheets": 1,
            }
        ],
        "source_ids": ["src_01"],
        "variant_tags": [],
        "quality_flags": [],
    }


def test_llm_runtime_extracts_verified_row_and_unresolved_others(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)

    result = run_llm_multi_agent_extraction(
        _manifest(workbook_path),
        llm_client=_FakeLLMClient(),
    )

    rows = result["rows"]
    assert len(rows) == len(STARTER_CONCEPT_IDS)

    ebitda_row = next(row for row in rows if row["concept_id"] == "ebitda_reported")
    assert ebitda_row["status"] == "verified"
    assert ebitda_row["extraction_reason_code"] is None
    assert ebitda_row["uncertainty_source"] is None
    assert ebitda_row["source_modality"] == "table_cell"
    assert ebitda_row["evidence"]["doc_id"] == "file_01"
    assert ebitda_row["evidence"]["locator_value"] == "A2"
    assert "llm_trace" in ebitda_row
    assert ebitda_row["llm_trace"]["attempts"]
    assert ebitda_row["llm_trace"]["final_reason"] == "accepted"

    unresolved = [row for row in rows if row["concept_id"] != "ebitda_reported"]
    assert all(row["status"] == "unresolved" for row in unresolved)
    assert all("no_reliable_candidate" in row["hard_blockers"] for row in unresolved)
    assert all("llm_trace" in row for row in unresolved)
    assert all(
        row["extraction_reason_code"] in {None, "current_package_missing_exact_support", "multiple_matching_rows"}
        for row in unresolved
    )
    assert all(
        (
            row["extraction_reason_code"] is None
            and row["uncertainty_source"] is None
        )
        or row["uncertainty_source"] == "package_extraction"
        for row in unresolved
    )


def test_process_package_manifest_supports_llm_mode(monkeypatch, tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)
    manifest = _manifest(workbook_path)

    def _fake_run_llm(package_manifest: dict, **kwargs):
        return run_llm_multi_agent_extraction(package_manifest, llm_client=_FakeLLMClient())

    monkeypatch.setattr("agent_app_dataset.internal_processing.run_llm_multi_agent_extraction", _fake_run_llm)

    payload, summary = process_package_manifest(
        package_manifest=manifest,
        labels_dir=None,
        events_log_path=tmp_path / "events.jsonl",
        extraction_mode="llm",
        max_retries=2,
    )

    assert payload["packages"][0]["package_id"] == "pkg_llm_001"
    assert summary["rows"] == len(STARTER_CONCEPT_IDS)


def test_llm_runtime_handles_missing_sources_with_unresolved_rows() -> None:
    manifest = {
        "package_id": "pkg_llm_002",
        "workspace_id": "ws_default",
        "deal_id": "deal_llm",
        "period_end_date": "2026-02-28",
        "source_email_id": "email_llm_002",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_02",
                "source_id": "src_02",
                "doc_type": "PDF",
                "filename": "missing.pdf",
                "storage_uri": "s3://missing/missing.pdf",
                "checksum": "checksum_02",
                "pages_or_sheets": 1,
            }
        ],
    }

    result = run_llm_multi_agent_extraction(manifest, llm_client=_StubLLMClient())
    assert len(result["rows"]) == len(STARTER_CONCEPT_IDS)
    assert all(row["status"] == "unresolved" for row in result["rows"])
    assert any("missing_source_document" in row["hard_blockers"] for row in result["rows"])
    assert all(row["extraction_reason_code"] == "current_package_missing_exact_support" for row in result["rows"])
    assert all(row["uncertainty_source"] == "package_extraction" for row in result["rows"])
    assert all(row["expected_section_state"] == "source_document_unavailable" for row in result["rows"])


def test_llm_runtime_captures_exact_row_header_missing_reason(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)

    result = run_llm_multi_agent_extraction(
        _manifest(workbook_path),
        llm_client=_FakeLLMUnanchoredClient(),
    )

    ebitda_row = next(row for row in result["rows"] if row["concept_id"] == "ebitda_reported")
    assert ebitda_row["status"] == "candidate_flagged"
    assert ebitda_row["extraction_reason_code"] == "exact_row_header_missing"
    assert ebitda_row["extraction_reason_label"] == "Exact row header missing"
    assert ebitda_row["uncertainty_source"] == "package_extraction"
    assert ebitda_row["source_modality"] == "pdf_text"
    assert ebitda_row["candidate_count"] >= 1


def test_llm_runtime_captures_pdf_text_only_reason(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)

    result = run_llm_multi_agent_extraction(
        _manifest(workbook_path),
        llm_client=_FakeLLMPdfCandidateClient(),
    )

    ebitda_row = next(row for row in result["rows"] if row["concept_id"] == "ebitda_reported")
    assert ebitda_row["status"] == "candidate_flagged"
    assert ebitda_row["extraction_reason_code"] == "candidate_from_pdf_text_only"
    assert ebitda_row["extraction_reason_label"] == "Extracted from PDF text only"
    assert ebitda_row["uncertainty_source"] == "package_extraction"
    assert ebitda_row["source_modality"] == "pdf_text"
