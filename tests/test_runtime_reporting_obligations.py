from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from agent_app_dataset.internal_api import create_app


class _FixedObligationLLMClient:
    def __init__(self, candidates: list[dict]) -> None:
        self._candidates = candidates
        self.model_name = "test-obligation-llm"

    def run_json(self, *, agent_id: str, system_prompt: str, user_payload: dict) -> dict:
        assert agent_id == "agent_reporting_obligation_candidates"
        return {"candidates": self._candidates}


def _write_borrower_sheet(path: Path, rows: list[tuple[str, float]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage"
    for idx, (label, value) in enumerate(rows, start=1):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
    wb.save(path)


def _write_requirement_sheet(path: Path, lines: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    for idx, line in enumerate(lines, start=1):
        ws[f"A{idx}"] = line
    wb.save(path)


def _ingest_payload(
    *,
    deal_id: str,
    source_email_id: str,
    period_end_date: str,
    received_at: str,
    file_id: str,
    source_id: str,
    path: Path,
) -> dict:
    return {
        "sender_email": "ops@borrower.com",
        "source_email_id": source_email_id,
        "deal_id": deal_id,
        "period_end_date": period_end_date,
        "received_at": received_at,
        "files": [
            {
                "file_id": file_id,
                "source_id": source_id,
                "doc_type": "XLSX",
                "filename": path.name,
                "storage_uri": str(path),
                "checksum": f"checksum_{source_id}",
                "pages_or_sheets": 1,
            }
        ],
        "variant_tags": ["runtime_test"],
        "quality_flags": [],
    }


def _process_runtime(client: TestClient, package_id: str) -> None:
    response = client.post(
        f"/internal/v1/packages/{package_id}:process",
        json={"async_mode": False, "max_retries": 1, "extraction_mode": "runtime"},
    )
    assert response.status_code == 200
    assert response.json()["status"] in {"completed", "needs_review"}


def _seed_baseline_and_current(client: TestClient, tmp_path: Path, deal_id: str) -> tuple[str, str]:
    baseline_path = tmp_path / f"{deal_id}_baseline.xlsx"
    current_path = tmp_path / f"{deal_id}_current.xlsx"
    _write_borrower_sheet(
        baseline_path,
        [
            ("Revenue total", 12_500_000),
            ("Net Income", 980_000),
            ("EBITDA adjusted", 2_610_000),
        ],
    )
    _write_borrower_sheet(
        current_path,
        [
            ("Revenue total", 12_300_000),
            ("EBITDA adjusted", 2_450_000),
            ("Interest expense", 460_000),
        ],
    )

    baseline_ingest = client.post(
        "/internal/v1/packages:ingest",
        json=_ingest_payload(
            deal_id=deal_id,
            source_email_id=f"{deal_id}_baseline",
            period_end_date="2025-06-30",
            received_at="2025-07-05T12:00:00+00:00",
            file_id=f"{deal_id}_file_baseline",
            source_id=f"{deal_id}_src_baseline",
            path=baseline_path,
        ),
    )
    current_ingest = client.post(
        "/internal/v1/packages:ingest",
        json=_ingest_payload(
            deal_id=deal_id,
            source_email_id=f"{deal_id}_current",
            period_end_date="2025-09-30",
            received_at="2025-10-05T12:00:00+00:00",
            file_id=f"{deal_id}_file_current",
            source_id=f"{deal_id}_src_current",
            path=current_path,
        ),
    )
    assert baseline_ingest.status_code == 200
    assert current_ingest.status_code == 200
    return baseline_ingest.json()["package_id"], current_ingest.json()["package_id"]


def test_runtime_ingests_reporting_obligation_catalog(tmp_path: Path) -> None:
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    deal_create = client.post(
        "/internal/v1/deals",
        json={"display_name": "Runtime Obligation Deal", "deal_id": "deal_runtime_obligation"},
    )
    assert deal_create.status_code == 200
    deal_id = deal_create.json()["deal_id"]

    requirement_path = tmp_path / "reporting_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        [
            "Borrower shall provide Net Income with each quarterly reporting package.",
        ],
    )

    ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_001",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_checksum_001",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "off",
        },
    )
    assert ingest.status_code == 200
    payload = ingest.json()
    assert payload["grounded_extracted"] >= 1

    listed = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligations")
    assert listed.status_code == 200
    obligations = listed.json()["obligations"]
    assert obligations
    grounded = [row for row in obligations if row["grounding_state"] == "grounded"]
    assert grounded
    grounded_net_income = [row for row in grounded if row["required_concept_id"] == "net_income"]
    assert grounded_net_income
    obligation = grounded_net_income[0]
    assert obligation["doc_id"] == "req_doc_001"
    assert obligation["doc_name"] == requirement_path.name
    assert obligation["locator_type"]
    assert obligation["locator_value"]
    assert obligation["source_snippet"]
    assert obligation["grounding_state"] == "grounded"


def test_create_deal_setup_can_ingest_reporting_requirement_docs(tmp_path: Path) -> None:
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    deal_id = "deal_setup_requirement_ingest"
    requirement_path = tmp_path / "setup_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        [
            "Borrower shall provide Net Income with each quarterly reporting package.",
        ],
    )

    deal_create = client.post(
        "/internal/v1/deals",
        json={
            "display_name": "Setup Requirement Deal",
            "deal_id": deal_id,
            "reporting_requirement_docs": [
                {
                    "doc_id": "req_doc_setup_001",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_setup_checksum_001",
                    "pages_or_sheets": 1,
                }
            ],
            "reporting_requirement_llm_discovery": "off",
        },
    )
    assert deal_create.status_code == 200
    payload = deal_create.json()
    setup = payload["reporting_requirement_setup"]
    assert setup["mode"] == "ingested_during_deal_setup"
    assert setup["docs_received"] == 1
    assert setup["grounded_extracted"] >= 1

    listed = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligations")
    assert listed.status_code == 200
    grounded = [row for row in listed.json()["obligations"] if row["grounding_state"] == "grounded"]
    assert grounded
    assert any(row["required_concept_id"] == "net_income" for row in grounded)

    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)
    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)
    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    net_income = next(item for item in queue.json()["items"] if item["metric_key"] == "net_income")
    assert net_income["case_mode"] == "investigation_missing_required_reporting"
    assert net_income["obligation_grounding_state"] == "grounded"


def test_runtime_review_queue_builds_requirement_anchor_from_catalog(tmp_path: Path) -> None:
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    deal_id = "deal_runtime_req_anchor"
    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)

    requirement_path = tmp_path / "runtime_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        ["Borrower must include Net Income in each quarterly reporting package."],
    )
    obligation_ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_runtime_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_runtime_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "off",
        },
    )
    assert obligation_ingest.status_code == 200
    assert obligation_ingest.json()["grounded_extracted"] >= 1

    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)

    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    by_metric = {item["metric_key"]: item for item in queue.json()["items"]}
    assert "net_income" in by_metric
    assert "ebitda_adjusted" in by_metric
    net_income = by_metric["net_income"]
    ebitda_adjusted = by_metric["ebitda_adjusted"]

    assert net_income["case_mode"] == "investigation_missing_required_reporting"
    assert net_income["obligation_grounding_state"] == "grounded"
    assert net_income["concept_maturity"] == "grounded"
    assert net_income["proof_compare_mode"] == "baseline_current_plus_requirement"
    requirement = net_income["requirement_anchor"]
    assert requirement["doc_id"] == "req_doc_runtime_01"
    assert requirement["required_concept_id"] == "net_income"
    assert requirement["doc_name"] == requirement_path.name
    assert requirement["locator_type"]
    assert requirement["locator_value"]
    assert requirement["grounded"] is True
    assert requirement["source_snippet"]
    assert requirement["preview_url"].endswith("/reporting-obligations/" + requirement["obligation_id"] + "/preview")
    assert requirement["download_url"].endswith("/document:download")
    assert net_income["baseline_anchor"] is not None
    assert net_income["baseline_anchor"]["value_display"] not in {"", "N/A"}
    assert net_income["current_search_state"] in {"missing", "candidate_only", "candidate_unanchored"}
    assert net_income["grounded_implication"] == "Required support for Net Income is missing from the current package."

    action_labels = {entry["label"] for entry in net_income["recommended_actions"]}
    assert "Request borrower update" in action_labels
    assert "Confirm alternate source" in action_labels
    assert "Mark item received" in action_labels
    assert "View reporting requirement" in action_labels
    assert "Copy borrower draft" in action_labels
    assert net_income["primary_action"]["label"] == "Request borrower update"

    draft = net_income["draft_borrower_query"]
    assert isinstance(draft, dict)
    assert "subject" in draft and "body" in draft and "text" in draft
    assert "Net Income" in draft["subject"]
    assert "deal_runtime_req_anchor" in draft["body"]
    assert ("Sep 2025" in draft["body"]) or ("2025-09-30" in draft["body"])
    assert requirement_path.name in draft["body"]
    assert "Please provide the authoritative support for Net Income" in draft["body"]

    assert ebitda_adjusted["concept_maturity"] == "review"
    assert ebitda_adjusted["trust_tier"] == "review"
    assert ebitda_adjusted["authority_level"] == "analyst_confirmation_required"
    assert ebitda_adjusted["review_required"] is True
    assert ebitda_adjusted["case_mode"] == "review_possible_material_change"
    assert "confirm source evidence" in ebitda_adjusted["grounded_implication"].lower()

    delta = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/delta")
    assert delta.status_code == 200
    net_income_row = next(row for row in delta.json()["rows"] if row["concept_id"] == "net_income")
    trace = client.get(f"/internal/v1/traces/{net_income_row['trace_id']}")
    assert trace.status_code == 200
    trace_row = trace.json()["row"]
    # Runtime requirement anchor is catalog-driven in review queue, not eval-row pass-through.
    assert trace_row.get("requirement_anchor") is None


def test_runtime_falls_back_when_no_grounded_requirement_exists(tmp_path: Path) -> None:
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    deal_id = "deal_runtime_no_requirement"
    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)
    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)

    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    by_metric = {item["metric_key"]: item for item in queue.json()["items"]}
    net_income = by_metric["net_income"]
    assert net_income["case_mode"] in {"investigation_missing_source", "investigation_candidate_only"}
    assert net_income["case_mode"] != "investigation_missing_required_reporting"
    assert net_income["obligation_grounding_state"] == "not_grounded"
    assert net_income["requirement_anchor"] is None
    assert net_income["proof_compare_mode"] in {"baseline_vs_current_missing", "baseline_vs_current_candidate"}
    assert "Requirement anchor:" not in net_income["grounded_implication"]


def test_ambiguous_requirement_text_does_not_trigger_grounded_case_mode(tmp_path: Path) -> None:
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    deal_id = "deal_runtime_ambiguous_requirement"
    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)

    requirement_path = tmp_path / "ambiguous_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        [
            "Net Income appears in the reporting package index.",
            "Quarterly reporting package overview.",
        ],
    )
    ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_ambiguous_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_ambiguous_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "off",
        },
    )
    assert ingest.status_code == 200
    assert ingest.json()["grounded_extracted"] == 0

    listed = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligations")
    assert listed.status_code == 200
    obligations = listed.json()["obligations"]
    assert obligations
    assert all(row["grounding_state"] != "grounded" for row in obligations)

    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)

    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    net_income = next(item for item in queue.json()["items"] if item["metric_key"] == "net_income")
    assert net_income["case_mode"] in {"investigation_missing_source", "investigation_candidate_only"}
    assert net_income["case_mode"] != "investigation_missing_required_reporting"
    assert net_income["obligation_grounding_state"] == "not_grounded"


def test_llm_candidate_promotes_to_grounded_and_drives_queue_case(tmp_path: Path) -> None:
    llm_client = _FixedObligationLLMClient(
        candidates=[
            {
                "doc_id": "req_doc_llm_01",
                "locator_type": "cell",
                "locator_value": "A1",
                "source_snippet": "Borrower shall provide net profit with each quarterly reporting package.",
                "candidate_obligation_type": "reporting_requirement",
                "candidate_concept_id": "net_income",
                "reason": "net profit maps to net income requirement",
                "certainty_bucket": "high",
            }
        ]
    )
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
        reporting_obligation_llm_client=llm_client,
    )
    client = TestClient(app)

    deal_id = "deal_runtime_llm_promoted"
    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)

    requirement_path = tmp_path / "llm_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        ["Borrower shall provide net profit with each quarterly reporting package."],
    )
    ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_llm_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_llm_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "on",
        },
    )
    assert ingest.status_code == 200
    payload = ingest.json()
    assert payload["deterministic_obligations_extracted"] == 0
    assert payload["candidate_discovery"]["status"] == "completed"
    assert payload["candidate_discovery"]["summary"]["promoted"] >= 1
    assert payload["grounded_extracted"] >= 1

    candidates = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligation-candidates")
    assert candidates.status_code == 200
    candidate_rows = candidates.json()["candidates"]
    assert candidate_rows
    promoted = [row for row in candidate_rows if row["candidate_concept_id"] == "net_income"]
    assert promoted
    assert promoted[0]["grounding_state"] == "grounded"
    assert promoted[0]["promoted"] is True
    assert promoted[0]["promoted_obligation_id"]
    assert promoted[0]["model_name"] == "test-obligation-llm"

    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)

    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    net_income = next(item for item in queue.json()["items"] if item["metric_key"] == "net_income")
    assert net_income["case_mode"] == "investigation_missing_required_reporting"
    assert net_income["obligation_grounding_state"] == "grounded"
    assert net_income["requirement_anchor"] is not None
    assert net_income["requirement_anchor"]["doc_name"] == requirement_path.name
    assert net_income["requirement_anchor"]["required_concept_id"] == "net_income"


def test_llm_candidate_not_promoted_does_not_drive_missing_required_case(tmp_path: Path) -> None:
    llm_client = _FixedObligationLLMClient(
        candidates=[
            {
                "doc_id": "req_doc_llm_amb_01",
                "locator_type": "cell",
                "locator_value": "A1",
                "source_snippet": "Borrower provides financial information quarterly.",
                "candidate_obligation_type": "reporting_requirement",
                "candidate_concept_id": "net_income",
                "reason": "possible reporting reference",
                "certainty_bucket": "medium",
            }
        ]
    )
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
        reporting_obligation_llm_client=llm_client,
    )
    client = TestClient(app)

    deal_id = "deal_runtime_llm_unpromoted"
    baseline_pkg, current_pkg = _seed_baseline_and_current(client, tmp_path, deal_id)

    requirement_path = tmp_path / "llm_ambiguous_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        ["Borrower provides financial information quarterly."],
    )
    ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_llm_amb_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_llm_amb_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "on",
        },
    )
    assert ingest.status_code == 200
    payload = ingest.json()
    assert payload["candidate_discovery"]["status"] == "completed"
    assert payload["candidate_discovery"]["summary"]["grounded"] == 0
    assert payload["candidate_discovery"]["summary"]["promoted"] == 0
    assert payload["grounded_extracted"] == 0

    candidates = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligation-candidates")
    assert candidates.status_code == 200
    candidate_rows = candidates.json()["candidates"]
    assert candidate_rows
    assert all(row["promoted"] is False for row in candidate_rows)
    assert all(row["promoted_obligation_id"] in {None, ""} for row in candidate_rows)
    assert all(row["grounding_state"] in {"ambiguous", "unsupported"} for row in candidate_rows)

    _process_runtime(client, baseline_pkg)
    _process_runtime(client, current_pkg)
    queue = client.get(f"/internal/v1/deals/{deal_id}/periods/{current_pkg}/review_queue")
    assert queue.status_code == 200
    net_income = next(item for item in queue.json()["items"] if item["metric_key"] == "net_income")
    assert net_income["case_mode"] in {"investigation_missing_source", "investigation_candidate_only"}
    assert net_income["case_mode"] != "investigation_missing_required_reporting"
    assert net_income["requirement_anchor"] is None
    assert net_income["obligation_grounding_state"] == "not_grounded"


def test_ebitda_requirement_candidate_remains_non_grounded(tmp_path: Path) -> None:
    llm_client = _FixedObligationLLMClient(
        candidates=[
            {
                "doc_id": "req_doc_ebitda_01",
                "locator_type": "cell",
                "locator_value": "A1",
                "source_snippet": "Borrower shall provide EBITDA with each quarterly reporting package.",
                "candidate_obligation_type": "reporting_requirement",
                "candidate_concept_id": "ebitda_reported",
                "reason": "explicit EBITDA requirement candidate",
                "certainty_bucket": "high",
            }
        ]
    )
    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
        reporting_obligation_llm_client=llm_client,
    )
    client = TestClient(app)

    deal_id = "deal_runtime_ebitda_candidate_only"
    deal_create = client.post(
        "/internal/v1/deals",
        json={"display_name": "EBITDA Candidate Deal", "deal_id": deal_id},
    )
    assert deal_create.status_code == 200

    requirement_path = tmp_path / "ebitda_requirements.xlsx"
    _write_requirement_sheet(
        requirement_path,
        ["Borrower shall provide EBITDA with each quarterly reporting package."],
    )
    ingest = client.post(
        f"/internal/v1/deals/{deal_id}/reporting-obligations:ingest",
        json={
            "docs": [
                {
                    "doc_id": "req_doc_ebitda_01",
                    "doc_type": "XLSX",
                    "filename": requirement_path.name,
                    "storage_uri": str(requirement_path),
                    "checksum": "req_ebitda_checksum_01",
                    "pages_or_sheets": 1,
                }
            ],
            "clear_existing_for_docs": True,
            "llm_discovery": "on",
        },
    )
    assert ingest.status_code == 200
    payload = ingest.json()
    assert payload["candidate_discovery"]["status"] == "completed"
    assert payload["candidate_discovery"]["summary"]["grounded"] == 0
    assert payload["candidate_discovery"]["summary"]["promoted"] == 0

    candidates = client.get(
        f"/internal/v1/deals/{deal_id}/reporting-obligation-candidates",
        params={"candidate_concept_id": "ebitda_reported"},
    )
    assert candidates.status_code == 200
    candidate_rows = candidates.json()["candidates"]
    assert candidate_rows
    assert all(row["candidate_concept_id"] == "ebitda_reported" for row in candidate_rows)
    assert all(row["grounding_state"] in {"ambiguous", "unsupported"} for row in candidate_rows)
    assert all(row["promoted"] is False for row in candidate_rows)

    obligations = client.get(f"/internal/v1/deals/{deal_id}/reporting-obligations")
    assert obligations.status_code == 200
    assert all(row["required_concept_id"] != "ebitda_reported" for row in obligations.json()["obligations"])
