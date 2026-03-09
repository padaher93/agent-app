from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from agent_app_dataset.internal_api import create_app
from agent_app_dataset.runtime_extractor import runtime_extract_package_predictions


def _build_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws["A1"] = "Revenue total"
    ws["B1"] = 1250
    ws["A2"] = "Adjusted EBITDA"
    ws["B2"] = 300
    ws["A3"] = "Interest expense"
    ws["B3"] = 60
    ws["A4"] = "Total debt"
    ws["B4"] = 820
    wb.save(path)


def _build_ambiguous_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws["A1"] = "Adjusted EBITDA pending verification"
    ws["A2"] = "Adjusted EBITDA pending verification"
    ws["A3"] = "Revenue total"
    ws["B3"] = 1250
    wb.save(path)


def test_runtime_extractor_reads_xlsx_without_labels(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)

    package_manifest = {
        "schema_version": "1.0",
        "package_id": "pkg_runtime_0001",
        "deal_id": "deal_runtime",
        "period_end_date": "2026-01-31",
        "source_email_id": "email_runtime_0001",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_runtime_01",
                "source_id": "src_runtime_01",
                "doc_type": "XLSX",
                "filename": "borrower_update.xlsx",
                "storage_uri": str(workbook_path),
                "checksum": "runtime_checksum_01",
                "pages_or_sheets": 1,
            }
        ],
        "source_ids": ["src_runtime_01"],
        "variant_tags": [],
        "quality_flags": [],
        "labeling_workflow": {
            "primary_labeler_status": "not_started",
            "reviewer_status": "not_started",
            "adjudication_status": "not_required",
        },
    }

    prediction = runtime_extract_package_predictions(package_manifest)
    rows = prediction["rows"]
    revenue = next(row for row in rows if row["concept_id"] == "revenue_total")
    unresolved = next(row for row in rows if row["concept_id"] == "total_assets")

    assert revenue["status"] in {"verified", "candidate_flagged"}
    assert revenue["normalized_value"] == 1250.0
    assert revenue["evidence"]["doc_id"] == "file_runtime_01"
    assert revenue["evidence"]["locator_type"] == "cell"
    assert unresolved["evidence"]["doc_id"] == "file_runtime_01"
    assert unresolved["evidence"]["locator_value"] != ""


def test_runtime_extractor_persists_multiple_matching_rows_reason(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update_ambiguous.xlsx"
    _build_ambiguous_xlsx(workbook_path)

    package_manifest = {
        "schema_version": "1.0",
        "package_id": "pkg_runtime_ambiguous",
        "deal_id": "deal_runtime",
        "period_end_date": "2026-01-31",
        "source_email_id": "email_runtime_ambiguous",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_runtime_ambiguous",
                "source_id": "src_runtime_ambiguous",
                "doc_type": "XLSX",
                "filename": "borrower_update_ambiguous.xlsx",
                "storage_uri": str(workbook_path),
                "checksum": "runtime_checksum_ambiguous",
                "pages_or_sheets": 1,
            }
        ],
        "source_ids": ["src_runtime_ambiguous"],
        "variant_tags": [],
        "quality_flags": [],
        "labeling_workflow": {
            "primary_labeler_status": "not_started",
            "reviewer_status": "not_started",
            "adjudication_status": "not_required",
        },
    }

    prediction = runtime_extract_package_predictions(package_manifest)
    ebitda_adjusted = next(row for row in prediction["rows"] if row["concept_id"] == "ebitda_adjusted")

    assert ebitda_adjusted["status"] in {"candidate_flagged", "unresolved"}
    assert ebitda_adjusted["extraction_reason_code"] == "multiple_matching_rows"
    assert ebitda_adjusted["extraction_reason_label"] == "Multiple matching rows"
    assert ebitda_adjusted["uncertainty_source"] == "package_extraction"
    assert ebitda_adjusted["candidate_count"] >= 2


def test_internal_api_runtime_mode_processes_without_label_files(tmp_path: Path) -> None:
    workbook_path = tmp_path / "borrower_update.xlsx"
    _build_sample_xlsx(workbook_path)

    app = create_app(
        db_path=tmp_path / "runtime" / "api.sqlite3",
        labels_dir=tmp_path / "labels",
        events_log_path=tmp_path / "runtime" / "events.jsonl",
    )
    client = TestClient(app)

    ingest_payload = {
        "sender_email": "ops@borrower.com",
        "source_email_id": "email_runtime_0002",
        "deal_id": "deal_runtime",
        "period_end_date": "2026-01-31",
        "received_at": "2026-03-06T12:00:00+00:00",
        "files": [
            {
                "file_id": "file_runtime_02",
                "source_id": "src_runtime_02",
                "doc_type": "XLSX",
                "filename": "borrower_update.xlsx",
                "storage_uri": str(workbook_path),
                "checksum": "runtime_checksum_02",
                "pages_or_sheets": 1,
            }
        ],
        "variant_tags": ["runtime_test"],
        "quality_flags": [],
    }

    ingest = client.post("/internal/v1/packages:ingest", json=ingest_payload)
    assert ingest.status_code == 200
    package_id = ingest.json()["package_id"]

    process = client.post(
        f"/internal/v1/packages/{package_id}:process",
        json={"async_mode": False, "max_retries": 2, "extraction_mode": "runtime"},
    )
    assert process.status_code == 200
    assert process.json()["status"] in {"completed", "needs_review"}

    delta = client.get(f"/internal/v1/deals/deal_runtime/periods/{package_id}/delta")
    assert delta.status_code == 200
    rows = delta.json()["rows"]
    revenue = next(row for row in rows if row["concept_id"] == "revenue_total")
    assert revenue["current_value"] == 1250.0

    evidence = client.get(f"/internal/v1/traces/{revenue['trace_id']}/evidence")
    assert evidence.status_code == 200
    preview = evidence.json()["evidence_preview"]["preview"]
    assert preview["kind"] == "xlsx_sheet"
